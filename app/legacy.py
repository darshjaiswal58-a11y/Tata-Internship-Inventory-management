from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from email.parser import BytesParser
from email.policy import default as email_policy
from pathlib import Path
from urllib.parse import parse_qs, urlparse
import io
import json
import mimetypes
import sqlite3
import sys
import uuid
from datetime import datetime
from threading import Lock

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
VENDOR_DIR = BASE_DIR / "vendor"

try:
    import bcrypt
except ModuleNotFoundError:
    if VENDOR_DIR.exists():
        sys.path.insert(0, str(VENDOR_DIR))
        import bcrypt
    else:
        raise

PACKAGED_APP = (BASE_DIR / "PACKAGED_APP").exists()
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
REPORT_DIR = DATA_DIR / "reports"
ANALYSIS_DIR = DATA_DIR / "analysis"
INVENTORY_ANALYSIS_DIR = DATA_DIR / "inventory_analysis"
CRITERIA_FILE = DATA_DIR / "criteria.json"
GROUP_CRITERIA_FILE = DATA_DIR / "group_criteria.json"
STOCK_OVERRIDES_FILE = DATA_DIR / "stock_overrides.json"
USERS_DB = DATA_DIR / "users.db"
MATERIAL_CRITICAL_DEFAULTS_FILE = DATA_DIR / "material_critical_defaults.json"
STOCK_ZONE_ANALYSIS_FILE = DATA_DIR / "stock_zone_analysis.json"
UPLOADS_FILE = DATA_DIR / "uploads.json"
INVENTORY_UPLOADS_FILE = DATA_DIR / "inventory_uploads.json"
MATERIALS_FILE = DATA_DIR / "materials.json"
MATERIAL_GROUPS_FILE = DATA_DIR / "material_groups.json"
MATERIAL_GROUP_SOURCE = BASE_DIR.parent / "updated" / "Material group 3002 and 3004.xlsx"
MATERIAL_GROUP_NAMES_SOURCE = Path("E:/downloads/Material_Group_Withno name.xlsx")
MATERIAL_GROUP_MISSING_NAMES_SOURCE = Path("E:/downloads/material names not provided for these (1).xlsx")
MATERIAL_GROUP_NAMES_SOURCES = [
    MATERIAL_GROUP_NAMES_SOURCE,
    MATERIAL_GROUP_MISSING_NAMES_SOURCE,
]
MATERIAL_CRITICAL_DEFAULTS_SOURCE = Path("E:/downloads/Phase1_Critical_Stock_Analysis.xlsx")
STOCK_PARTS_SOURCE = BASE_DIR.parent / "updated" / "criteria matched" / "180_Eligible_Materials_Updated_Usage.xlsx"
STOCK_ZONE_ANALYSIS_SOURCE = Path("E:/downloads/330_Eligible_Materials_Zone_Classification.xlsx")
TOOLS_CRITICAL_DEFAULTS_SOURCE = BASE_DIR.parent / "updated" / "stockitemlist30043002plant" / "Machinery_Spares_Critical_Values_Final.xlsx"
TOOLS_STOCK_LIST_SOURCE = BASE_DIR.parent / "updated" / "stockitemlist30043002plant" / "final stock list 25-26 (MACHINERY SPARES) jan dark room.xlsx"

REQUIRED_COLUMNS = [
    "Purchase Order Date",
    "Entry Date",
    "Material",
    "Material Description",
    "Quantity",
    "Days Between",
    "Valuated Stock",
]

INITIAL_USERS = ()

SESSIONS = {}
MATERIAL_GROUP_CACHE = None
MATERIAL_CRITICAL_DEFAULTS_CACHE = None
STOCK_ZONE_ANALYSIS_CACHE = None
UPLOAD_PROGRESS = {}
UPLOAD_PROGRESS_LOCK = Lock()

CATEGORIES = {
    "ml_spare": "Machinery Spare",
    "tools": "Tools",
    "plant_3003": "Plant 3003",
    "plant_3005": "Plant 3005",
}
DEFAULT_CATEGORY = "ml_spare"
RED_ZONE_FILTER = "__red_zone__"
MATERIAL_CRITICAL_DEFAULT_SOURCES = {
    DEFAULT_CATEGORY: [MATERIAL_CRITICAL_DEFAULTS_SOURCE],
    "tools": [TOOLS_STOCK_LIST_SOURCE, TOOLS_CRITICAL_DEFAULTS_SOURCE],
}
MATERIAL_CRITICAL_DEFAULTS_VERSION = 3
STOCK_ZONE_ANALYSIS_VERSION = 9
STOCK_PLANT_BY_CATEGORY = {
    DEFAULT_CATEGORY: "3002",
    "tools": "3004",
    "plant_3003": "3003",
    "plant_3005": "3005",
}
STOCK_ZONE_CATEGORIES = {DEFAULT_CATEGORY, "tools"}


def ensure_dirs():
    for path in [STATIC_DIR, DATA_DIR, UPLOAD_DIR, REPORT_DIR, ANALYSIS_DIR, INVENTORY_ANALYSIS_DIR]:
        path.mkdir(parents=True, exist_ok=True)
    for file_path, empty in [
        (CRITERIA_FILE, {}),
        (GROUP_CRITERIA_FILE, {}),
        (STOCK_OVERRIDES_FILE, {"keep": {}, "remove": {}}),
        (MATERIAL_CRITICAL_DEFAULTS_FILE, {"materials": {}}),
        (STOCK_ZONE_ANALYSIS_FILE, {"counts": {"red": 0, "yellow": 0, "green": 0}, "materials": {}}),
        (UPLOADS_FILE, []),
        (INVENTORY_UPLOADS_FILE, []),
        (MATERIALS_FILE, {}),
        (MATERIAL_GROUPS_FILE, {"groups": {}, "materials": {}}),
    ]:
        if not file_path.exists():
            file_path.write_text(json.dumps(empty, indent=2), encoding="utf-8")
    init_user_db()
    if not PACKAGED_APP:
        sync_material_groups()
        sync_material_critical_defaults()
        sync_stock_zone_analysis()


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path, payload):
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def set_upload_progress(job_id, percent, message, state="processing", category=None):
    if not job_id:
        return
    percent = max(0, min(100, int(percent)))
    payload = {
        "job_id": job_id,
        "percent": percent,
        "message": message,
        "state": state,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    if category:
        payload["category"] = category
        payload["category_label"] = CATEGORIES.get(category, category)
    with UPLOAD_PROGRESS_LOCK:
        existing = UPLOAD_PROGRESS.get(job_id, {})
        UPLOAD_PROGRESS[job_id] = {**existing, **payload}


def get_upload_progress(job_id):
    if not job_id:
        return {"state": "missing", "percent": 0, "message": "Waiting for upload to start."}
    with UPLOAD_PROGRESS_LOCK:
        progress = dict(UPLOAD_PROGRESS.get(job_id, {}))
    if not progress:
        return {"job_id": job_id, "state": "pending", "percent": 0, "message": "Waiting for upload to start."}
    return progress


def db_connection():
    connection = sqlite3.connect(USERS_DB)
    connection.row_factory = sqlite3.Row
    return connection


def user_from_row(row):
    if not row:
        return None
    return {
        "id": row["id"],
        "name": row["name"],
        "email": row["email"],
        "role": row["role"],
    }


def get_user_by_email(email):
    email = str(email or "").strip().lower()
    if not email:
        return None
    with db_connection() as connection:
        row = connection.execute(
            "SELECT id, name, email, password_hash, role FROM users WHERE email = ?",
            (email,),
        ).fetchone()
    return dict(row) if row else None


def create_user(name, email, password, role):
    email = str(email or "").strip().lower()
    if not email or not password:
        raise ValueError("Email and password are required.")
    if get_user_by_email(email) is not None:
        raise ValueError("Email already registered.")
    password_hash = bcrypt.hashpw(str(password).encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    with db_connection() as connection:
        connection.execute(
            "INSERT INTO users (name, email, password_hash, role) VALUES (?, ?, ?, ?)",
            (str(name or "").strip(), email, password_hash, str(role or "employee").strip()),
        )
    return user_from_row(get_user_by_email(email))


def authenticate_user(email, password):
    user = get_user_by_email(email)
    if not user:
        return None
    password_hash = str(user.get("password_hash", "")).encode("utf-8")
    if not bcrypt.checkpw(str(password or "").encode("utf-8"), password_hash):
        return None
    return user_from_row(user)


def create_session(email):
    email = str(email or "").strip().lower()
    session_id = uuid.uuid4().hex
    SESSIONS[session_id] = email
    with db_connection() as connection:
        connection.execute(
            "INSERT OR REPLACE INTO sessions (id, email, created_at) VALUES (?, ?, ?)",
            (session_id, email, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        )
    return session_id


def email_for_session(session_id):
    if not session_id:
        return None
    email = SESSIONS.get(session_id)
    if email:
        return email
    with db_connection() as connection:
        row = connection.execute(
            "SELECT email FROM sessions WHERE id = ?",
            (session_id,),
        ).fetchone()
    if not row:
        return None
    email = row["email"]
    SESSIONS[session_id] = email
    return email


def delete_session(session_id):
    if not session_id:
        return
    SESSIONS.pop(session_id, None)
    with db_connection() as connection:
        connection.execute("DELETE FROM sessions WHERE id = ?", (session_id,))


def init_user_db():
    with db_connection() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY,
                name TEXT,
                email TEXT UNIQUE,
                password_hash TEXT,
                role TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS sessions (
                id TEXT PRIMARY KEY,
                email TEXT,
                created_at TEXT
            )
            """
        )
    for name, email, password, role in INITIAL_USERS:
        try:
            create_user(name, email, password, role)
        except ValueError:
            pass


def sync_material_groups():
    global MATERIAL_GROUP_CACHE
    if PACKAGED_APP and MATERIAL_GROUPS_FILE.exists():
        MATERIAL_GROUP_CACHE = read_json(MATERIAL_GROUPS_FILE)
        return
    if not MATERIAL_GROUP_SOURCE.exists():
        return
    source_mtime = MATERIAL_GROUP_SOURCE.stat().st_mtime
    names_mtimes = {
        str(source): source.stat().st_mtime
        for source in MATERIAL_GROUP_NAMES_SOURCES
        if source.exists()
    }
    existing = read_json(MATERIAL_GROUPS_FILE) if MATERIAL_GROUPS_FILE.exists() else {}
    if existing.get("source_mtime") == source_mtime and existing.get("group_names_mtimes") == names_mtimes:
        MATERIAL_GROUP_CACHE = existing
        return

    wb = load_workbook(MATERIAL_GROUP_SOURCE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    try:
        material_idx = headers.index("Material")
        group_idx = headers.index("Material Group")
    except ValueError:
        return
    description_idx = headers.index("Short Text") if "Short Text" in headers else None

    groups = {}
    material_lookup = {}
    for row in rows:
        material = str(row[material_idx] or "").strip()
        material_group = str(row[group_idx] or "").strip()
        if not material or not material_group:
            continue
        description = str(row[description_idx] or "").strip() if description_idx is not None else ""
        groups.setdefault(material_group, {
            "code": material_group,
            "label": material_group,
            "source_parts_count": 0,
        })
        groups[material_group]["source_parts_count"] += 1
        material_lookup[material] = {
            "material": material,
            "material_group": material_group,
            "description": description,
        }

    payload = {
        "source_file": str(MATERIAL_GROUP_SOURCE),
        "source_mtime": source_mtime,
        "group_names_files": [str(source) for source in MATERIAL_GROUP_NAMES_SOURCES if source.exists()],
        "group_names_mtimes": names_mtimes,
        "groups": dict(sorted(groups.items())),
        "materials": material_lookup,
    }
    apply_material_group_labels(payload)
    write_json(MATERIAL_GROUPS_FILE, payload)
    MATERIAL_GROUP_CACHE = payload


def load_material_group_labels():
    labels = {}
    for source in MATERIAL_GROUP_NAMES_SOURCES:
        if not source.exists():
            continue
        wb = load_workbook(source, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        group_idx = find_column(headers, ["Material Group"])
        description_idx = find_column(headers, ["Material Group Description", "Description", "Group Description"])
        if group_idx is None or description_idx is None:
            continue
        for row in rows:
            group = str(row[group_idx] or "").strip() if group_idx < len(row) else ""
            description = str(row[description_idx] or "").strip() if description_idx < len(row) and row[description_idx] is not None else ""
            if group and description and description.upper() != "#N/A":
                labels[group] = description
    return labels


def apply_material_group_labels(payload):
    labels = load_material_group_labels()
    for code, group in payload.get("groups", {}).items():
        group["label"] = labels.get(code, group.get("label") or code)


def sync_material_critical_defaults():
    global MATERIAL_CRITICAL_DEFAULTS_CACHE
    if PACKAGED_APP and MATERIAL_CRITICAL_DEFAULTS_FILE.exists():
        MATERIAL_CRITICAL_DEFAULTS_CACHE = read_json(MATERIAL_CRITICAL_DEFAULTS_FILE)
        return
    source_mtimes = {
        f"{category}:{source.name}": source.stat().st_mtime
        for category, sources in MATERIAL_CRITICAL_DEFAULT_SOURCES.items()
        for source in sources
        if source.exists()
    }
    if not source_mtimes:
        return
    existing = read_json(MATERIAL_CRITICAL_DEFAULTS_FILE) if MATERIAL_CRITICAL_DEFAULTS_FILE.exists() else {}
    if existing.get("source_mtimes") == source_mtimes and existing.get("loader_version") == MATERIAL_CRITICAL_DEFAULTS_VERSION:
        MATERIAL_CRITICAL_DEFAULTS_CACHE = existing
        return

    categories = {}
    combined_defaults = {}
    skipped_without_material = 0
    source_files = {}
    for category, sources in MATERIAL_CRITICAL_DEFAULT_SOURCES.items():
        category_defaults = {}
        source_files[category] = []
        for source in sources:
            if not source.exists():
                continue
            wb = load_workbook(source, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
            material_idx = find_column(headers, ["Material", "Part Number", "Part No"])
            group_idx = find_column(headers, ["Material Category", "Material Group", "Category"])
            description_idx = find_column(headers, ["Description", "Material Description", "Material_Description", "Short Text", "Part Description", "Part Desc"])
            critical_idx = find_column(headers, ["Critical Value", "Estimated_Critical_Value", "Critical Stock", "Minimum Stock", "Min"])
            consumption_idx = find_column(headers, ["Net_Consumption", "Net Consumption", "Total_Consumption", "CONSUMPTION", "Consumption", "Usage Frequency", "Usage_Frequency"])
            current_stock_idx = find_column(headers, ["Current Stock", "Stock 23.04.2026", "Stock 09.12.2025"])
            oldest_issue_idx = find_column(headers, ["Oldest_Issue_Date", "Oldest Issue Date"])
            if material_idx is None or critical_idx is None:
                continue
            for row in rows:
                material = str(row[material_idx] or "").strip() if material_idx < len(row) else ""
                if not material:
                    skipped_without_material += 1
                    continue
                material_group = str(row[group_idx] or "").strip() if group_idx is not None and group_idx < len(row) else ""
                description = str(row[description_idx] or "").strip() if description_idx is not None and description_idx < len(row) and row[description_idx] is not None else ""
                critical_value = abs(parse_number(row[critical_idx])) if critical_idx < len(row) else 0
                net_consumption = abs(parse_number(row[consumption_idx])) if consumption_idx is not None and consumption_idx < len(row) else 0
                current_stock = parse_number(row[current_stock_idx], "") if current_stock_idx is not None and current_stock_idx < len(row) else ""
                oldest_issue_date = normalize_date(row[oldest_issue_idx]) if oldest_issue_idx is not None and oldest_issue_idx < len(row) else ""
                default = {
                    "material": material,
                    "category": category,
                    "material_group": material_group,
                    "description": description,
                    "critical_value": critical_value,
                    "minimum_stock": critical_value,
                    "net_consumption": net_consumption,
                    "current_stock": current_stock,
                    "oldest_issue_date": oldest_issue_date,
                }
                category_defaults[material] = default
                combined_defaults[material] = default
                if material_group:
                    learn_material_group(material, material_group, description, persist=False)
            source_files[category].append(str(source))
        categories[category] = {"source_files": source_files[category], "materials": category_defaults}

    payload = {
        "source_files": source_files,
        "source_mtimes": source_mtimes,
        "loader_version": MATERIAL_CRITICAL_DEFAULTS_VERSION,
        "categories": categories,
        "materials": combined_defaults,
        "skipped_without_material": skipped_without_material,
        "loaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    write_json(MATERIAL_CRITICAL_DEFAULTS_FILE, payload)
    save_material_group_data(get_material_group_data())
    MATERIAL_CRITICAL_DEFAULTS_CACHE = payload


def get_material_critical_defaults():
    global MATERIAL_CRITICAL_DEFAULTS_CACHE
    if MATERIAL_CRITICAL_DEFAULTS_CACHE is None:
        sync_material_critical_defaults()
    if MATERIAL_CRITICAL_DEFAULTS_CACHE is None:
        MATERIAL_CRITICAL_DEFAULTS_CACHE = read_json(MATERIAL_CRITICAL_DEFAULTS_FILE)
    return MATERIAL_CRITICAL_DEFAULTS_CACHE


def normalize_zone(value):
    value = str(value or "").strip().lower().replace("_", " ")
    if "not eligible" in value:
        return "not_eligible"
    if "red" in value:
        return "red"
    if "yellow" in value:
        return "yellow"
    if "green" in value:
        return "green"
    return ""


def count_excel_rows(path):
    if not path.exists():
        return 0
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    return sum(1 for row in rows if row and not all(value is None for value in row))


def load_stock_part_materials(path):
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    material_idx = find_column(headers, ["Material", "Part Number", "Part No"])
    plant_idx = find_column(headers, ["Plant"])
    description_idx = find_column(headers, ["Part Description", "Material Description", "Description"])
    usage_3m_idx = find_column(headers, ["Usage_Last_3_Months"])
    usage_6m_idx = find_column(headers, ["Usage_Last_6_Months"])
    if material_idx is None:
        return {}
    materials = {}
    for row in rows:
        if not row or material_idx >= len(row) or row[material_idx] is None:
            continue
        material = str(row[material_idx]).strip()
        if material:
            materials.setdefault(material, {
                "material": material,
                "plant": str(row[plant_idx]).strip() if plant_idx is not None and plant_idx < len(row) and row[plant_idx] is not None else "",
                "description": str(row[description_idx]).strip() if description_idx is not None and description_idx < len(row) and row[description_idx] is not None else "",
                "usage_last_3_months": parse_number(row[usage_3m_idx], "") if usage_3m_idx is not None and usage_3m_idx < len(row) else "",
                "usage_last_6_months": parse_number(row[usage_6m_idx], "") if usage_6m_idx is not None and usage_6m_idx < len(row) else "",
            })
    return materials


def load_stock_part_summary(path):
    summary = {
        "row_count": 0,
        "unique_count": 0,
        "row_counts_by_plant": {},
        "unique_counts_by_plant": {},
        "material_rows": {},
    }
    if not path.exists():
        return summary
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    material_idx = find_column(headers, ["Material", "Part Number", "Part No"])
    plant_idx = find_column(headers, ["Plant"])
    unique_materials = set()
    unique_by_plant = {}
    for row in rows:
        if not row or all(value is None for value in row):
            continue
        material = str(row[material_idx]).strip() if material_idx is not None and material_idx < len(row) and row[material_idx] is not None else ""
        plant = str(row[plant_idx]).strip() if plant_idx is not None and plant_idx < len(row) and row[plant_idx] is not None else ""
        summary["row_count"] += 1
        summary["row_counts_by_plant"][plant] = summary["row_counts_by_plant"].get(plant, 0) + 1
        if material:
            unique_materials.add(material)
            unique_by_plant.setdefault(plant, set()).add(material)
            detail = summary["material_rows"].setdefault(material, {"plant": plant, "rows": 0})
            detail["rows"] += 1
    summary["unique_count"] = len(unique_materials)
    summary["unique_counts_by_plant"] = {
        plant: len(materials)
        for plant, materials in unique_by_plant.items()
    }
    return summary


def kept_stock_overrides():
    overrides = read_json(STOCK_OVERRIDES_FILE)
    criteria = read_json(CRITERIA_FILE)
    kept = {}
    for item in overrides.get("keep", {}).values():
        material = str(item.get("material", "")).strip()
        category = normalize_category(item.get("category", DEFAULT_CATEGORY))
        if material:
            kept[material] = {
                "material": material,
                "plant": STOCK_PLANT_BY_CATEGORY.get(category, ""),
                "description": str(item.get("description", "")).strip(),
                "category": category,
                "material_group": str(item.get("material_group", "")).strip(),
                "usage_last_3_months": "",
                "usage_last_6_months": "",
                "keep_stock": True,
            }
    for item in criteria.values():
        if not item.get("keep_stock"):
            continue
        material = str(item.get("material", "")).strip()
        if not material:
            continue
        category = normalize_category(item.get("category", DEFAULT_CATEGORY))
        kept[material] = {
            "material": material,
            "plant": STOCK_PLANT_BY_CATEGORY.get(category, ""),
            "description": str(item.get("description", "")).strip(),
            "category": category,
            "material_group": material_group_for(material),
            "usage_last_3_months": "",
            "usage_last_6_months": "",
            "keep_stock": True,
        }
    return kept


def removed_stock_overrides():
    overrides = read_json(STOCK_OVERRIDES_FILE)
    return {
        str(item.get("material", "")).strip()
        for item in overrides.get("remove", {}).values()
        if str(item.get("material", "")).strip()
    }


def refresh_stock_after_override():
    global STOCK_ZONE_ANALYSIS_CACHE
    STOCK_ZONE_ANALYSIS_CACHE = None
    sync_stock_zone_analysis()


def apply_kept_stock_summary(summary, kept_records):
    for material, record in kept_records.items():
        plant = str(record.get("plant", "")).strip()
        summary["row_count"] += 1
        summary["unique_count"] += 1
        summary["row_counts_by_plant"][plant] = summary["row_counts_by_plant"].get(plant, 0) + 1
        summary["unique_counts_by_plant"][plant] = summary["unique_counts_by_plant"].get(plant, 0) + 1


def sync_stock_zone_analysis():
    global STOCK_ZONE_ANALYSIS_CACHE
    if PACKAGED_APP and STOCK_ZONE_ANALYSIS_FILE.exists():
        STOCK_ZONE_ANALYSIS_CACHE = read_json(STOCK_ZONE_ANALYSIS_FILE)
        return
    if not STOCK_ZONE_ANALYSIS_SOURCE.exists():
        return
    source_mtime = STOCK_ZONE_ANALYSIS_SOURCE.stat().st_mtime
    stock_parts_mtime = STOCK_PARTS_SOURCE.stat().st_mtime if STOCK_PARTS_SOURCE.exists() else None
    criteria_mtime = CRITERIA_FILE.stat().st_mtime if CRITERIA_FILE.exists() else None
    overrides_mtime = STOCK_OVERRIDES_FILE.stat().st_mtime if STOCK_OVERRIDES_FILE.exists() else None
    existing = read_json(STOCK_ZONE_ANALYSIS_FILE) if STOCK_ZONE_ANALYSIS_FILE.exists() else {}
    if (
        existing.get("source_mtime") == source_mtime
        and existing.get("stock_parts_mtime") == stock_parts_mtime
        and existing.get("criteria_mtime") == criteria_mtime
        and existing.get("overrides_mtime") == overrides_mtime
        and existing.get("loader_version") == STOCK_ZONE_ANALYSIS_VERSION
    ):
        STOCK_ZONE_ANALYSIS_CACHE = existing
        return

    wb = load_workbook(STOCK_ZONE_ANALYSIS_SOURCE, read_only=True, data_only=True)
    sheet_name = "All Materials" if "All Materials" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    material_idx = find_column(headers, ["Material", "Part Number", "Part No"])
    description_idx = find_column(headers, ["Material Description", "Description", "Part Description", "Part Desc"])
    group_idx = find_column(headers, ["Material Group", "Material Category", "Category", "Part Type"])
    eligible_idx = find_column(headers, ["Eligible"])
    current_stock_idx = find_column(headers, ["Current_Stock", "Current Stock", "Valuated Stock"])
    critical_idx = find_column(headers, ["Critical_Value", "Critical Value", "Estimated_Critical_Value", "Minimum Stock", "Min"])
    usage_idx = find_column(headers, ["Avg_Consumption", "Average Consumption", "Net_Consumption", "Net Consumption"])
    zone_idx = find_column(headers, ["Zone_Status", "Zone Status", "Stock Zone", "Zone"])

    counts = {"red": 0, "yellow": 0, "green": 0, "not_eligible": 0}
    materials = {}
    material_groups = {}
    skipped_without_zone = 0
    stock_part_records = load_stock_part_materials(STOCK_PARTS_SOURCE)
    kept_records = kept_stock_overrides()
    removed_records = removed_stock_overrides()
    stock_part_records = {
        material: record
        for material, record in stock_part_records.items()
        if material not in removed_records
    }
    stock_part_summary = load_stock_part_summary(STOCK_PARTS_SOURCE)
    for material in removed_records:
        detail = stock_part_summary.get("material_rows", {}).get(material)
        if not detail:
            continue
        plant = str(detail.get("plant", "")).strip()
        row_count = int(detail.get("rows", 1) or 1)
        summary_count = stock_part_summary["row_counts_by_plant"].get(plant, 0)
        if summary_count > 0:
            stock_part_summary["row_counts_by_plant"][plant] = max(0, summary_count - row_count)
            stock_part_summary["row_count"] = max(0, stock_part_summary["row_count"] - row_count)
        unique_count = stock_part_summary["unique_counts_by_plant"].get(plant, 0)
        if unique_count > 0:
            stock_part_summary["unique_counts_by_plant"][plant] = unique_count - 1
            stock_part_summary["unique_count"] = max(0, stock_part_summary["unique_count"] - 1)
    new_kept_records = {
        material: record
        for material, record in kept_records.items()
        if material not in stock_part_records and material not in removed_records
    }
    stock_part_records.update(new_kept_records)
    apply_kept_stock_summary(stock_part_summary, new_kept_records)
    stock_part_materials = set(stock_part_records)
    stock_parts_count = len(stock_part_records) or (count_excel_rows(STOCK_PARTS_SOURCE) if STOCK_PARTS_SOURCE.exists() else 0)
    if material_idx is None or zone_idx is None:
        payload = {
            "source_file": str(STOCK_ZONE_ANALYSIS_SOURCE),
            "source_mtime": source_mtime,
            "stock_parts_file": str(STOCK_PARTS_SOURCE) if STOCK_PARTS_SOURCE.exists() else "",
            "stock_parts_mtime": stock_parts_mtime,
            "criteria_mtime": criteria_mtime,
            "overrides_mtime": overrides_mtime,
            "loader_version": STOCK_ZONE_ANALYSIS_VERSION,
            "sheet": sheet_name,
            "counts": counts,
            "total_materials": 0,
            "materials": {},
            "error": "Missing Material or Zone_Status column.",
            "loaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        write_json(STOCK_ZONE_ANALYSIS_FILE, payload)
        STOCK_ZONE_ANALYSIS_CACHE = payload
        return

    for row in rows:
        if not row or all(value is None for value in row):
            continue
        material = str(row[material_idx] or "").strip() if material_idx < len(row) else ""
        if material in removed_records:
            continue
        if stock_part_materials and material not in stock_part_materials:
            continue
        zone = normalize_zone(row[zone_idx] if zone_idx < len(row) else "")
        material_group = str(row[group_idx] or "").strip() if group_idx is not None and group_idx < len(row) and row[group_idx] is not None else ""
        if not material_group and material:
            material_group = material_group_for(material)
        eligible = str(row[eligible_idx] or "").strip().lower() == "yes" if eligible_idx is not None and eligible_idx < len(row) else zone != "not_eligible"
        if not material:
            continue
        if not zone:
            skipped_without_zone += 1
            continue
        counts[zone] += 1
        if material_group:
            group_counts = material_groups.setdefault(material_group, {
                "total": 0,
                "eligible": 0,
                "not_eligible": 0,
                "red": 0,
                "yellow": 0,
                "green": 0,
            })
            group_counts["total"] += 1
            if eligible:
                group_counts["eligible"] += 1
            else:
                group_counts["not_eligible"] += 1
            if zone in ["red", "yellow", "green"]:
                group_counts[zone] += 1
        stock_part = stock_part_records.get(material, {})
        materials[material] = {
            "material": material,
            "description": stock_part.get("description") or (str(row[description_idx] or "").strip() if description_idx is not None and description_idx < len(row) and row[description_idx] is not None else ""),
            "material_group": material_group,
            "plant": stock_part.get("plant", ""),
            "current_stock": parse_number(row[current_stock_idx], "") if current_stock_idx is not None and current_stock_idx < len(row) else "",
            "critical_value": parse_number(row[critical_idx], "") if critical_idx is not None and critical_idx < len(row) else "",
            "avg_consumption": stock_part.get("usage_last_6_months", "") if stock_part.get("usage_last_6_months", "") != "" else (parse_number(row[usage_idx], "") if usage_idx is not None and usage_idx < len(row) else ""),
            "usage_last_3_months": stock_part.get("usage_last_3_months", ""),
            "usage_last_6_months": stock_part.get("usage_last_6_months", ""),
            "eligible": eligible,
            "zone": zone,
            "zone_label": "Not Eligible" if zone == "not_eligible" else f"{zone.title()} Zone",
        }

    for material, stock_part in new_kept_records.items():
        if material in materials:
            continue
        material_group = stock_part.get("material_group") or material_group_for(material)
        counts["green"] += 1
        if material_group:
            group_counts = material_groups.setdefault(material_group, {
                "total": 0,
                "eligible": 0,
                "not_eligible": 0,
                "red": 0,
                "yellow": 0,
                "green": 0,
            })
            group_counts["total"] += 1
            group_counts["eligible"] += 1
            group_counts["green"] += 1
        materials[material] = {
            "material": material,
            "description": stock_part.get("description", ""),
            "material_group": material_group,
            "plant": stock_part.get("plant", ""),
            "current_stock": "",
            "critical_value": "",
            "avg_consumption": "",
            "usage_last_3_months": "",
            "usage_last_6_months": "",
            "eligible": True,
            "zone": "green",
            "zone_label": "Manual Keep",
            "keep_stock": True,
        }

    eligible_count = counts["red"] + counts["yellow"] + counts["green"]
    for group_counts in material_groups.values():
        group_counts["uploaded"] = group_counts["total"]
        group_counts["criteria"] = group_counts["eligible"]
        group_counts["active"] = group_counts["eligible"]
        group_counts["critical"] = group_counts["red"] + group_counts["yellow"]

    payload = {
        "source_file": str(STOCK_ZONE_ANALYSIS_SOURCE),
        "source_mtime": source_mtime,
        "stock_parts_file": str(STOCK_PARTS_SOURCE) if STOCK_PARTS_SOURCE.exists() else "",
        "stock_parts_mtime": stock_parts_mtime,
        "criteria_mtime": criteria_mtime,
        "overrides_mtime": overrides_mtime,
        "loader_version": STOCK_ZONE_ANALYSIS_VERSION,
        "sheet": sheet_name,
        "counts": counts,
        "total_materials": stock_parts_count or sum(counts.values()),
        "stock_parts_count": stock_parts_count or eligible_count,
        "eligible_materials": stock_parts_count or eligible_count,
        "critical_materials": counts["red"] + counts["yellow"],
        "stock_part_summary": stock_part_summary,
        "material_groups": material_groups,
        "materials": materials,
        "skipped_without_zone": skipped_without_zone,
        "loaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    write_json(STOCK_ZONE_ANALYSIS_FILE, payload)
    STOCK_ZONE_ANALYSIS_CACHE = payload


def get_stock_zone_analysis():
    global STOCK_ZONE_ANALYSIS_CACHE
    if STOCK_ZONE_ANALYSIS_CACHE is None:
        sync_stock_zone_analysis()
    if STOCK_ZONE_ANALYSIS_CACHE is None:
        STOCK_ZONE_ANALYSIS_CACHE = read_json(STOCK_ZONE_ANALYSIS_FILE)
    return STOCK_ZONE_ANALYSIS_CACHE


def stock_zone_analysis_for_category(category):
    category = normalize_category(category)
    plant = STOCK_PLANT_BY_CATEGORY.get(category, "")
    source = get_stock_zone_analysis()
    if not plant:
        return source
    materials = {
        code: material
        for code, material in source.get("materials", {}).items()
        if str(material.get("plant", "")).strip() == plant
    }
    counts = {"red": 0, "yellow": 0, "green": 0, "not_eligible": 0}
    material_groups = {}
    for material in materials.values():
        zone = material.get("zone", "not_eligible")
        counts[zone] = counts.get(zone, 0) + 1
        group_code = material.get("material_group") or material_group_for(material.get("material"))
        if not group_code:
            continue
        group_counts = material_groups.setdefault(group_code, {
            "total": 0,
            "eligible": 0,
            "not_eligible": 0,
            "red": 0,
            "yellow": 0,
            "green": 0,
        })
        group_counts["total"] += 1
        if material.get("eligible", True):
            group_counts["eligible"] += 1
        else:
            group_counts["not_eligible"] += 1
        if zone in ["red", "yellow", "green"]:
            group_counts[zone] += 1
    for group_counts in material_groups.values():
        group_counts["uploaded"] = group_counts["total"]
        group_counts["criteria"] = group_counts["eligible"]
        group_counts["active"] = group_counts["eligible"]
        group_counts["critical"] = group_counts["red"] + group_counts["yellow"]
    eligible_count = counts["red"] + counts["yellow"] + counts["green"]
    stock_part_summary = source.get("stock_part_summary", {})
    plant_stock_parts = stock_part_summary.get("row_counts_by_plant", {}).get(plant, len(materials))
    return {
        **source,
        "category": category,
        "plant": plant,
        "counts": counts,
        "total_materials": len(materials),
        "stock_parts_count": plant_stock_parts,
        "eligible_materials": eligible_count,
        "critical_materials": counts["red"] + counts["yellow"],
        "material_groups": material_groups,
        "materials": materials,
    }


def default_critical_for(material, category=DEFAULT_CATEGORY):
    defaults = get_material_critical_defaults()
    category_defaults = defaults.get("categories", {}).get(normalize_category(category), {}).get("materials", {})
    return category_defaults.get(str(material), {})


def category_critical_defaults(category):
    return get_material_critical_defaults().get("categories", {}).get(normalize_category(category), {}).get("materials", {})


def get_category_materials(category):
    category = normalize_category(category)
    materials = {}
    for code, default in category_critical_defaults(category).items():
        materials[code] = {
            "material": code,
            "description": default.get("description", ""),
            "current_stock": default.get("current_stock", ""),
            "category": category,
            "material_group": default.get("material_group", ""),
            "critical_value": default.get("critical_value", ""),
            "net_consumption": default.get("net_consumption", ""),
            "last_entry_date": default.get("oldest_issue_date", ""),
            "last_quantity": "",
            "movement_type": "Default critical",
            "default_source": True,
        }
    for code, material in read_json(MATERIALS_FILE).items():
        if normalize_category(material.get("category", DEFAULT_CATEGORY)) != category:
            continue
        default = default_critical_for(code, category)
        materials[code] = {
            **materials.get(code, {}),
            **material,
            "critical_value": material.get("critical_value", default.get("critical_value", "")),
            "net_consumption": material.get("net_consumption", default.get("net_consumption", "")),
            "default_source": False,
        }
    return list(materials.values())


def get_material_group_data():
    global MATERIAL_GROUP_CACHE
    if MATERIAL_GROUP_CACHE is None:
        sync_material_groups()
    if MATERIAL_GROUP_CACHE is None:
        MATERIAL_GROUP_CACHE = read_json(MATERIAL_GROUPS_FILE)
    return MATERIAL_GROUP_CACHE


def material_group_for(material):
    material_data = get_material_group_data().get("materials", {}).get(str(material), {})
    return material_data.get("material_group", "")


def save_material_group_data(payload):
    global MATERIAL_GROUP_CACHE
    write_json(MATERIAL_GROUPS_FILE, payload)
    MATERIAL_GROUP_CACHE = payload


def learn_material_group(material, material_group, description="", persist=True):
    material = str(material or "").strip()
    material_group = str(material_group or "").strip()
    if not material or not material_group:
        return False
    group_data = get_material_group_data()
    groups = group_data.setdefault("groups", {})
    materials = group_data.setdefault("materials", {})

    if material_group not in groups:
        groups[material_group] = {
            "code": material_group,
            "label": material_group,
            "source_parts_count": 0,
        }

    previous_group = materials.get(material, {}).get("material_group")
    if previous_group != material_group:
        groups[material_group]["source_parts_count"] = parse_number(groups[material_group].get("source_parts_count")) + 1

    materials[material] = {
        "material": material,
        "material_group": material_group,
        "description": description or materials.get(material, {}).get("description", ""),
    }
    group_data["groups"] = dict(sorted(groups.items()))
    group_data["last_learned_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if persist:
        save_material_group_data(group_data)
    return True


def learn_material_groups_bulk(rows):
    changed = False
    for row in rows:
        changed = learn_material_group(
            row.get("material"),
            row.get("material_group"),
            row.get("description", ""),
            persist=False,
        ) or changed
    if changed:
        save_material_group_data(get_material_group_data())


def parse_number(value, default=0):
    if value is None or value == "":
        return default
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def normalize_header_key(value):
    return " ".join(str(value or "").replace("_", " ").strip().lower().split())


def find_column(headers, candidates):
    normalized_headers = {normalize_header_key(header): index for index, header in enumerate(headers)}
    for candidate in candidates:
        index = normalized_headers.get(normalize_header_key(candidate))
        if index is not None:
            return index
    return None


def parse_date_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def normalize_date(value):
    parsed = parse_date_value(value)
    if not parsed:
        return ""
    return parsed.strftime("%Y-%m-%d")


def normalize_category(value):
    value = str(value or "").strip().lower()
    return value if value in CATEGORIES else DEFAULT_CATEGORY


def criteria_key(category, material):
    return f"{normalize_category(category)}::{material}"


def group_criteria_key(category, material_group):
    return f"{normalize_category(category)}::{material_group}"


def criteria_matches(setting, category, material):
    setting_category = normalize_category(setting.get("category", DEFAULT_CATEGORY))
    return setting.get("material") == material and setting_category == normalize_category(category)


def get_material_criteria(criteria, category, material):
    category = normalize_category(category)
    direct = criteria.get(criteria_key(category, material))
    if direct:
        return direct
    legacy = criteria.get(material)
    if legacy and normalize_category(legacy.get("category", DEFAULT_CATEGORY)) == category:
        return legacy
    for setting in criteria.values():
        if criteria_matches(setting, category, material):
            return setting
    default = default_critical_for(material, category)
    if default:
        return {
            "material": str(material),
            "description": default.get("description", ""),
            "category": category,
            "minimum_stock": default.get("critical_value", 0),
            "reorder_quantity": default.get("net_consumption", 0),
            "net_consumption": default.get("net_consumption", 0),
            "material_group": default.get("material_group", ""),
            "default_critical": True,
            "active": True,
        }
    return None


def material_description_for(material):
    group_data = get_material_group_data().get("materials", {}).get(str(material), {})
    return group_data.get("description", "")


def analyze_stock_zone_materials(category, zone="red"):
    category = normalize_category(category)
    stock_zone_analysis = stock_zone_analysis_for_category(category)

    rows = []
    materials = stock_zone_analysis.get("materials", {})
    for material in materials.values():
        if material.get("zone") != zone:
            continue
        code = str(material.get("material", ""))
        critical_default = default_critical_for(code, category)
        current_stock = material.get("current_stock", "")
        critical_value = material.get("critical_value", critical_default.get("critical_value", ""))
        material_group = material.get("material_group") or material_group_for(code)
        rows.append({
            "material": code,
            "material_group": material_group,
            "description": material.get("description") or material_description_for(code),
            "current_stock": current_stock,
            "critical_value": critical_value,
            "net_consumption": material.get("avg_consumption", critical_default.get("net_consumption", "")),
            "last_quantity": "",
            "movement_type": material.get("zone_label", "Red Zone"),
            "last_entry_date": "",
            "reason": f"{material.get('zone_label', 'Red Zone')} from plant {stock_zone_analysis.get('plant', 'stock')} stock data",
        })

    rows.sort(key=lambda item: (parse_number(item["current_stock"], 0), item["material"]))
    return {
        "category": category,
        "material_group": RED_ZONE_FILTER,
        "zone": zone,
        "saved_criteria": {},
        "critical_stock": "",
        "total_parts": stock_zone_analysis.get("total_materials", len(materials)),
        "critical_count": len(rows),
        "examples": [],
        "rows": rows[:100],
    }


def analyze_group_critical(category, material_group, critical_stock=None):
    category = normalize_category(category)
    if material_group == RED_ZONE_FILTER:
        return analyze_stock_zone_materials(category, "red")

    requested_critical_stock = critical_stock
    if requested_critical_stock is None or requested_critical_stock == "":
        latest_upload, analyzed_rows = latest_analyzed_rows(category)
        _, failed_rows = latest_failed_rows(category)
        if latest_upload:
            group_rows = []
            critical_rows = []
            for row in analyzed_rows:
                group_code = row.get("material_group") or material_group_for(row.get("material"))
                if material_group and group_code != material_group:
                    continue
                group_rows.append(row)
            for row in failed_rows:
                group_code = row.get("material_group") or material_group_for(row.get("material"))
                if material_group and group_code != material_group:
                    continue
                critical_rows.append({
                    "material": row.get("material", ""),
                    "material_group": group_code,
                    "description": row.get("description", ""),
                    "current_stock": row.get("current_stock", ""),
                    "critical_value": row.get("minimum_stock", ""),
                    "net_consumption": row.get("net_consumption", ""),
                    "last_quantity": row.get("latest_quantity", ""),
                    "movement_type": row.get("movement_type", ""),
                    "last_entry_date": row.get("entry_date", ""),
                    "reason": row.get("reason", "Critical from latest uploaded Excel analysis"),
                })
            stocks = sorted(parse_number(row.get("current_stock"), None) for row in group_rows if parse_number(row.get("current_stock"), None) is not None)
            examples = []
            if stocks:
                examples = sorted(set([
                    stocks[0],
                    stocks[len(stocks) // 4],
                    stocks[len(stocks) // 2],
                    5,
                ]))
            critical_rows.sort(key=lambda item: (parse_number(item["current_stock"], 0), item["material"]))
            return {
                "category": category,
                "material_group": material_group,
                "source": "latest_upload_analysis",
                "upload": latest_upload,
                "saved_criteria": {},
                "critical_stock": "",
                "total_parts": len(group_rows),
                "critical_count": len(critical_rows),
                "examples": examples[:4],
                "rows": critical_rows[:100],
            }

    group_settings = read_json(GROUP_CRITERIA_FILE)
    saved = group_settings.get(group_criteria_key(category, material_group), {})
    use_material_defaults = critical_stock is None or critical_stock == ""
    if use_material_defaults:
        critical_stock = saved.get("critical_stock", "")
    use_material_defaults = critical_stock is None or critical_stock == ""
    critical_stock = parse_number(critical_stock, None)

    group_rows = []
    critical_rows = []
    for material in get_category_materials(category):
        code = str(material.get("material", ""))
        group_code = material.get("material_group") or material_group_for(code)
        if material_group and group_code != material_group:
            continue
        current_stock_value = material.get("current_stock", "")
        current_stock = parse_number(current_stock_value, None)
        critical_default = default_critical_for(code, category)
        row_critical_stock = parse_number(critical_default.get("critical_value"), None)
        row_net_consumption = critical_default.get("net_consumption", material.get("net_consumption", ""))
        description = material.get("description") or material_description_for(code)
        row = {
            "material": code,
            "material_group": group_code,
            "description": description,
            "current_stock": current_stock_value,
            "critical_value": row_critical_stock,
            "net_consumption": row_net_consumption,
            "last_quantity": material.get("last_quantity", 0),
            "movement_type": material.get("movement_type", ""),
            "last_entry_date": material.get("last_entry_date", ""),
        }
        group_rows.append(row)
        effective_critical = row_critical_stock if use_material_defaults else critical_stock
        if current_stock is not None and effective_critical is not None and current_stock <= effective_critical:
            critical_rows.append({
                **row,
                "reason": f"Current stock {current_stock:g} is at or below critical value {effective_critical:g}",
            })
        elif material.get("default_source") and effective_critical is not None:
            critical_rows.append({
                **row,
                "reason": f"Default critical material from Excel; critical value is {effective_critical:g}",
            })

    stocks = sorted(parse_number(row.get("current_stock"), None) for row in group_rows if parse_number(row.get("current_stock"), None) is not None)
    examples = []
    if stocks:
        examples = sorted(set([
            stocks[0],
            stocks[len(stocks) // 4],
            stocks[len(stocks) // 2],
            5,
        ]))
    critical_rows.sort(key=lambda item: (parse_number(item["current_stock"], 0), item["material"]))
    return {
        "category": category,
        "material_group": material_group,
        "saved_criteria": saved,
        "critical_stock": critical_stock,
        "total_parts": len(group_rows),
        "critical_count": len(critical_rows),
        "examples": examples[:4],
        "rows": critical_rows[:100],
    }


def zone_part_rows(category, zone, material_group=""):
    category = normalize_category(category)
    zone = normalize_zone(zone)
    if zone not in {"red", "yellow", "green"}:
        zone = "red"
    if category not in STOCK_ZONE_CATEGORIES:
        upload, analyzed_rows = latest_analyzed_rows(category)
        _, failed_rows = latest_failed_rows(category)
        failed_materials = {str(row.get("material", "")).strip() for row in failed_rows}
        if zone == "red":
            source_rows = failed_rows
            zone_label = "Red Zone"
        elif zone == "green":
            source_rows = [
                row for row in analyzed_rows
                if str(row.get("material", "")).strip() not in failed_materials
            ]
            zone_label = "Green Zone"
        else:
            source_rows = []
            zone_label = "Yellow Zone"
        rows = []
        for row in source_rows:
            group_code = row.get("material_group") or material_group_for(row.get("material"))
            if material_group and group_code != material_group:
                continue
            rows.append({
                "material": row.get("material", ""),
                "material_group": group_code,
                "description": row.get("description", ""),
                "zone": zone_label,
                "current_stock": row.get("current_stock", ""),
                "critical_value": row.get("critical_value", row.get("minimum_stock", "")),
                "net_consumption": row.get("net_consumption", ""),
                "plant": STOCK_PLANT_BY_CATEGORY.get(category, ""),
            })
        rows.sort(key=lambda item: (item["material_group"], item["material"]))
        return rows
    stock_zone_analysis = stock_zone_analysis_for_category(category)
    rows = []
    for material in stock_zone_analysis.get("materials", {}).values():
        group_code = material.get("material_group") or material_group_for(material.get("material"))
        if material_group and group_code != material_group:
            continue
        if material.get("zone") != zone:
            continue
        rows.append({
            "material": material.get("material", ""),
            "material_group": group_code,
            "description": material.get("description", ""),
            "zone": material.get("zone_label", zone.title()),
            "current_stock": material.get("current_stock", ""),
            "critical_value": material.get("critical_value", ""),
            "net_consumption": material.get("avg_consumption", ""),
            "plant": material.get("plant", stock_zone_analysis.get("plant", "")),
        })
    rows.sort(key=lambda item: (item["material_group"], item["material"]))
    return rows


def build_zone_export(category, zone, material_group=""):
    rows = zone_part_rows(category, zone, material_group)
    wb = Workbook()
    ws = wb.active
    ws.title = "Zone Parts"
    headers = ["Material", "Material Group", "Part Name", "Zone", "Current Stock", "Critical Value", "Net Consumption", "Plant"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row.get("material", ""),
            row.get("material_group", ""),
            row.get("description", ""),
            row.get("zone", ""),
            row.get("current_stock", ""),
            row.get("critical_value", ""),
            row.get("net_consumption", ""),
            row.get("plant", ""),
        ])
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 42)
    return wb


def json_response(handler, status, payload):
    body = json.dumps(payload, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def parse_multipart_form(handler):
    content_type = handler.headers.get("Content-Type", "")
    if not content_type.startswith("multipart/form-data"):
        return {"fields": {}, "files": {}}

    length = int(handler.headers.get("Content-Length", "0"))
    body = handler.rfile.read(length)
    message = BytesParser(policy=email_policy).parsebytes(
        b"Content-Type: " + content_type.encode("utf-8") + b"\r\n"
        b"MIME-Version: 1.0\r\n\r\n" + body
    )

    fields = {}
    files = {}
    for part in message.iter_parts():
        name = part.get_param("name", header="content-disposition")
        if not name:
            continue
        payload = part.get_payload(decode=True) or b""
        filename = part.get_filename()
        if filename:
            files[name] = {"filename": filename, "content": payload}
        else:
            charset = part.get_content_charset() or "utf-8"
            fields[name] = payload.decode(charset, errors="replace")
    return {"fields": fields, "files": files}


def parse_cookies(header):
    cookies = {}
    if not header:
        return cookies
    for part in header.split(";"):
        if "=" not in part:
            continue
        key, value = part.strip().split("=", 1)
        cookies[key] = value
    return cookies


def current_user(handler):
    cookies = parse_cookies(handler.headers.get("Cookie"))
    session_id = cookies.get("session_id")
    if not session_id:
        return None
    email = email_for_session(session_id)
    if not email:
        return None
    user = get_user_by_email(email)
    if not user:
        return None
    return user_from_row(user)


def require_user(handler):
    user = current_user(handler)
    if not user:
        json_response(handler, 401, {"error": "Please log in first."})
        return None
    return user


def require_role(handler, role):
    user = require_user(handler)
    if not user:
        return None
    if user["role"] != role:
        json_response(handler, 403, {"error": "You do not have permission for this action."})
        return None
    return user


def save_report(upload_id, failed_rows):
    report_path = REPORT_DIR / f"reorder_report_{upload_id}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Reorder Report"
    headers = [
        "Material",
        "Material Group",
        "Material Description",
        "Current Stock",
        "Minimum Stock",
        "Net Consumption",
        "Reorder Quantity",
        "Latest Movement",
        "Movement Type",
        "Entry Date",
        "Reason",
    ]
    ws.append(headers)
    for row in failed_rows:
        ws.append([
            row["material"],
            row.get("material_group", ""),
            row["description"],
            row["current_stock"],
            row["minimum_stock"],
            row.get("net_consumption", ""),
            row["reorder_quantity"],
            row["latest_quantity"],
            row["movement_type"],
            row["entry_date"],
            row["reason"],
        ])
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 42)
    wb.save(report_path)
    return report_path


def build_no_criteria_export(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "No Criteria Items"
    headers = [
        "Material",
        "Material Group",
        "Material Description",
        "Current Stock",
        "Net Consumption",
        "Latest Movement",
        "Movement Type",
        "Entry Date",
        "Purchase Order Date",
        "Days Between",
        "Reason",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row.get("material", ""),
            row.get("material_group", ""),
            row.get("description", ""),
            row.get("current_stock", ""),
            row.get("net_consumption", ""),
            row.get("latest_quantity", ""),
            row.get("movement_type", ""),
            row.get("entry_date", ""),
            row.get("purchase_order_date", ""),
            row.get("days_between", ""),
            row.get("reason", "No matching criteria found"),
        ])
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 42)
    return wb


def save_analysis(upload_id, failed_rows, no_criteria_count, analyzed_rows=None, no_criteria_rows=None):
    analysis_path = ANALYSIS_DIR / f"analysis_{upload_id}.json"
    payload = {
        "upload_id": upload_id,
        "analyzed_rows": analyzed_rows or [],
        "failed_rows": failed_rows,
        "no_criteria_count": no_criteria_count,
        "no_criteria_rows": no_criteria_rows or [],
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    write_json(analysis_path, payload)
    return analysis_path


def analysis_for_upload(upload):
    if not upload:
        return {}
    analysis_file = upload.get("analysis_file") or f"analysis_{upload.get('id', '')}.json"
    analysis_path = ANALYSIS_DIR / analysis_file
    if not analysis_path.exists():
        return {}
    return read_json(analysis_path)


def latest_upload_for_category(category):
    category = normalize_category(category)
    uploads = read_json(UPLOADS_FILE)
    return next((
        item for item in uploads
        if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
    ), None)


def latest_analysis_for_category(category):
    upload = latest_upload_for_category(category)
    return upload, analysis_for_upload(upload)


def latest_analyzed_rows(category):
    upload, analysis = latest_analysis_for_category(category)
    rows = analysis.get("analyzed_rows", [])
    if rows:
        return upload, rows
    if upload:
        return upload, get_category_materials(category)
    return None, []


def latest_failed_rows(category):
    upload, analysis = latest_analysis_for_category(category)
    return upload, analysis.get("failed_rows", [])


def upload_analysis_summary(category):
    upload, analyzed_rows = latest_analyzed_rows(category)
    _, failed_rows = latest_failed_rows(category)
    group_counts = {}
    for row in analyzed_rows:
        group_code = row.get("material_group") or material_group_for(row.get("material"))
        if not group_code:
            continue
        counts = group_counts.setdefault(group_code, {
            "uploaded": 0,
            "active": 0,
            "critical": 0,
        })
        counts["uploaded"] += 1
        counts["active"] += 1
    for row in failed_rows:
        group_code = row.get("material_group") or material_group_for(row.get("material"))
        if group_code:
            group_counts.setdefault(group_code, {
                "uploaded": 0,
                "active": 0,
                "critical": 0,
            })["critical"] += 1
    return {
        "upload": upload,
        "uploaded": len(analyzed_rows),
        "active": len(analyzed_rows),
        "critical": len(failed_rows),
        "groups": group_counts,
        "failed_rows": failed_rows,
    }


def process_upload(file_path, original_name, category=DEFAULT_CATEGORY, uploaded_by="", progress_callback=None):
    category = normalize_category(category)
    criteria = read_json(CRITERIA_FILE)
    materials = read_json(MATERIALS_FILE)
    uploads = read_json(UPLOADS_FILE)

    def progress(percent, message):
        if progress_callback:
            progress_callback(percent, message)

    progress(8, "Opening Excel workbook.")
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Invalid Excel file: {exc}")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    first_data_index = next((index for index, row in enumerate(all_rows) if row and not all(value is None or value == "" for value in row)), None)
    if first_data_index is None:
        raise ValueError("Excel sheet is empty.")

    progress(14, "Reading Excel headers.")
    header_row = all_rows[first_data_index]
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    idx = {
        "Purchase Order Date": find_column(headers, ["Purchase Order Date", "PO Date", "Document Date", "Purchase Doc Date"]),
        "Entry Date": find_column(headers, ["Entry Date", "Posting Date", "GR Date", "Goods Receipt Date"]),
        "Material": find_column(headers, ["Material", "Part Number", "Part No", "Material Code"]),
        "Material Description": find_column(headers, ["Material Description", "Material Desc", "Description", "Part Description", "Part Desc", "Short Text", "Text"]),
        "Quantity": find_column(headers, ["Quantity", "Qty in unit of entry", "Qty in order unit", "Qty in OPUn", "Receipt Quantity", "Qty"]),
        "Days Between": find_column(headers, ["Days Between", "Days_Between", "Ageing Days", "Aging Days", "Lead Time Days"]),
        "Valuated Stock": find_column(headers, ["Valuated Stock", "Current Stock", "Current_Stock", "Stock", "Stock Qty", "Unrestricted Stock", "Qty in unit of entry", "Quantity"]),
        "Movement Type": find_column(headers, ["Movement Type", "Movement type", "MvT", "Movement"]),
        "Plant": find_column(headers, ["Plant"]),
    }
    data_start_index = first_data_index + 1
    material_group_idx = find_column(headers, [
        "Material Group",
        "Material Category",
        "Part Type",
        "Category",
    ])
    headerless_sap_dump = False
    if idx["Material"] is None and len(header_row) >= 19 and str(header_row[1] or "").strip() in {"3003", "3005"}:
        headerless_sap_dump = True
        idx.update({
            "Material": 0,
            "Plant": 1,
            "Material Description": 2,
            "Movement Type": 4,
            "Purchase Order Date": 9,
            "Entry Date": 17,
            "Quantity": 10,
            "Valuated Stock": 10,
            "Days Between": 18,
        })
        material_group_idx = 3
        data_start_index = first_data_index
    if idx["Valuated Stock"] is None:
        idx["Valuated Stock"] = idx["Quantity"]
    required_labels = ["Material", "Quantity", "Valuated Stock"]
    missing = [label for label in required_labels if idx.get(label) is None]
    if idx["Days Between"] is None and (idx["Purchase Order Date"] is None or idx["Entry Date"] is None):
        missing.append("Days Between or both Entry/Posting Date and Purchase Order/Document Date")
    if missing:
        raise ValueError(
            "Missing required columns: "
            + ", ".join(missing)
            + ". Accepted alternatives include SAP headers like Posting Date, Document Date, Qty in unit of entry, and Quantity."
        )
    progress(18, f"Analyzing {CATEGORIES.get(category, category)} rows.")
    row_iterator = all_rows[data_start_index:]
    total_data_rows = max(len(row_iterator), 1)
    category_plant = STOCK_PLANT_BY_CATEGORY.get(category, "")
    upload_id = uuid.uuid4().hex[:10]
    imported_rows = 0
    received_rows = 0
    used_rows = 0
    used_quantity_total = 0
    no_criteria_count = 0
    failed_by_material = {}
    analyzed_by_material = {}
    no_criteria_by_material = {}
    learned_material_groups = {}
    criteria_lookup = {
        str(setting.get("material", "")).strip(): setting
        for setting in criteria.values()
        if str(setting.get("material", "")).strip()
        and normalize_category(setting.get("category", DEFAULT_CATEGORY)) == category
    }

    errors = []

    def _is_number(value):
        if value is None or value == "":
            return False
        if isinstance(value, (int, float)):
            return True
        try:
            # allow numbers with commas
            float(str(value).replace(",", "").strip())
            return True
        except Exception:
            return False

    def _validate_date_field(value):
        if value is None or value == "":
            return True
        return parse_date_value(value) is not None

    for offset, raw in enumerate(row_iterator, start=1):
        row_index = data_start_index + offset
        processed_rows = offset
        if processed_rows == 1 or processed_rows % 50 == 0 or processed_rows >= total_data_rows:
            row_percent = 18 + min(67, int((processed_rows / total_data_rows) * 67))
            progress(row_percent, f"Analyzed {min(processed_rows, total_data_rows)} of {total_data_rows} rows.")
        if not raw or all(value is None for value in raw):
            continue

        row_errors = []

        def cell(label):
            column_index = idx.get(label)
            if column_index is None or column_index >= len(raw):
                return None
            return raw[column_index]

        material_cell = cell("Material")
        if material_cell is None:
            errors.append(f"Row {row_index}: missing 'Material' column value")
            continue

        material = str(material_cell or "").strip()
        if not material:
            continue
        plant_cell = cell("Plant")
        row_plant = str(plant_cell or "").strip()
        if category_plant and row_plant and row_plant != category_plant:
            continue

        description = str(cell("Material Description") or "").strip()
        uploaded_material_group = ""
        if material_group_idx is not None and material_group_idx < len(raw):
            uploaded_material_group = str(raw[material_group_idx] or "").strip()
        material_group = uploaded_material_group or material_group_for(material)
        critical_default = default_critical_for(material, category)
        if not description:
            description = str(critical_default.get("description") or materials.get(material, {}).get("description", "")).strip()
        net_consumption = critical_default.get("net_consumption", "")
        if uploaded_material_group:
            learned_material_groups[material] = {
                "material": material,
                "material_group": uploaded_material_group,
                "description": description,
            }
        # Validate numeric fields
        q_cell = cell("Quantity")
        vs_cell = cell("Valuated Stock")
        db_cell = cell("Days Between")
        ed_cell = cell("Entry Date")
        pod_cell = cell("Purchase Order Date")
        movement_type_cell = cell("Movement Type")

        if not _is_number(q_cell):
            row_errors.append(f"Row {row_index}: 'Quantity' must be numeric (got: {q_cell!r})")
        if not _is_number(vs_cell):
            row_errors.append(f"Row {row_index}: 'Valuated Stock/Stock Quantity' must be numeric (got: {vs_cell!r})")
        # Days Between may be empty for some sheets, but if present must be numeric
        if db_cell is not None and db_cell != "" and not _is_number(db_cell):
            if not (headerless_sap_dump and str(db_cell).strip().upper().startswith("Z-")):
                row_errors.append(f"Row {row_index}: 'Days Between' must be numeric (got: {db_cell!r})")

        # Validate date fields (if present)
        if not _validate_date_field(ed_cell):
            row_errors.append(f"Row {row_index}: 'Entry/Posting Date' not parseable (got: {ed_cell!r})")
        if not _validate_date_field(pod_cell):
            row_errors.append(f"Row {row_index}: 'Purchase Order/Document Date' not parseable (got: {pod_cell!r})")

        # If any validation errors collected so far, skip further processing for this row
        if row_errors:
            errors.extend(row_errors)
            # continue scanning to collect multiple errors, but don't process business logic
            continue

        quantity = parse_number(q_cell)
        current_stock = parse_number(vs_cell)
        entry_date_value = parse_date_value(ed_cell)
        purchase_order_date_value = parse_date_value(pod_cell)
        if db_cell is None or db_cell == "":
            days_between = (entry_date_value - purchase_order_date_value).days if entry_date_value and purchase_order_date_value else 0
        else:
            days_between = 0 if headerless_sap_dump and str(db_cell).strip().upper().startswith("Z-") else parse_number(db_cell)
        entry_date = normalize_date(ed_cell)
        purchase_order_date = normalize_date(pod_cell)
        movement_code = str(movement_type_cell or "").strip()
        is_used_movement = movement_code in {"201", "221", "261", "281", "543", "551", "601", "901"} or movement_code.startswith("9")
        is_received_movement = movement_code in {"101", "105", "501", "561"}
        movement_type = "Used" if is_used_movement or quantity < 0 else "Received" if is_received_movement or quantity > 0 else "No movement"
        if movement_type == "Received":
            received_rows += 1
        elif movement_type == "Used":
            used_rows += 1
            used_quantity_total += abs(quantity)

        materials[material] = {
            "material": material,
            "description": description,
            "current_stock": current_stock,
            "category": category,
            "material_group": material_group,
            "critical_value": critical_default.get("critical_value", ""),
            "net_consumption": net_consumption,
            "last_entry_date": entry_date,
            "last_quantity": quantity,
            "movement_type": movement_type,
            "days_between": days_between,
            "purchase_order_date": purchase_order_date,
        }
        analyzed_by_material[material] = {
            "material": material,
            "material_group": material_group,
            "description": description,
            "current_stock": current_stock,
            "critical_value": critical_default.get("critical_value", ""),
            "net_consumption": net_consumption,
            "latest_quantity": quantity,
            "movement_type": movement_type,
            "entry_date": entry_date,
            "purchase_order_date": purchase_order_date,
        }
        imported_rows += 1

        material_criteria = criteria_lookup.get(material)
        if material_criteria is None and critical_default:
            material_criteria = {
                "material": material,
                "description": critical_default.get("description", ""),
                "category": category,
                "minimum_stock": critical_default.get("critical_value", 0),
                "reorder_quantity": critical_default.get("net_consumption", 0),
                "net_consumption": critical_default.get("net_consumption", 0),
                "material_group": critical_default.get("material_group", ""),
                "default_critical": True,
                "active": True,
            }
        if not material_criteria or not material_criteria.get("active", True):
            no_criteria_count += 1
            no_criteria_by_material[material] = {
                "material": material,
                "material_group": material_group,
                "description": description,
                "current_stock": current_stock,
                "net_consumption": net_consumption,
                "latest_quantity": quantity,
                "movement_type": movement_type,
                "entry_date": entry_date,
                "purchase_order_date": purchase_order_date,
                "days_between": days_between,
                "reason": "No active matching criteria found for this material",
            }
            continue

        minimum_stock = parse_number(material_criteria.get("minimum_stock"))
        reorder_quantity = parse_number(material_criteria.get("reorder_quantity"))
        net_consumption = material_criteria.get("net_consumption", net_consumption)
        if current_stock <= minimum_stock:
            failed_by_material[material] = {
                "material": material,
                "material_group": material_group,
                "description": description,
                "current_stock": current_stock,
                "minimum_stock": minimum_stock,
                "net_consumption": net_consumption,
                "reorder_quantity": reorder_quantity,
                "latest_quantity": quantity,
                "movement_type": movement_type,
                "entry_date": entry_date,
                "reason": f"Current stock {current_stock:g} is at or below minimum stock {minimum_stock:g}",
            }

    if learned_material_groups:
        progress(87, "Updating material group lookup.")
        learn_material_groups_bulk(learned_material_groups.values())

    # If any validation errors were collected, reject the upload with a clear message.
    if errors:
        # limit output to first 20 errors for brevity
        summary = "; ".join(errors[:20])
        more = f"; and {len(errors)-20} more errors" if len(errors) > 20 else ""
        raise ValueError("Upload validation failed: " + summary + more)

    failed_rows = sorted(failed_by_material.values(), key=lambda item: (item["current_stock"], item["material"]))
    analyzed_rows = sorted(analyzed_by_material.values(), key=lambda item: item["material"])
    no_criteria_rows = sorted(no_criteria_by_material.values(), key=lambda item: item["material"])
    progress(90, "Building Excel reports.")
    report_path = save_report(upload_id, failed_rows)
    analysis_path = save_analysis(upload_id, failed_rows, no_criteria_count, analyzed_rows, no_criteria_rows)
    upload_record = {
        "id": upload_id,
        "file_name": original_name,
        "category": category,
        "category_label": CATEGORIES[category],
        "uploaded_by": uploaded_by,
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "imported_rows": imported_rows,
        "received_rows": received_rows,
        "used_rows": used_rows,
        "used_quantity_total": used_quantity_total,
        "material_group_column_found": material_group_idx is not None,
        "learned_material_groups": len({
            row["material_group"]
            for row in learned_material_groups.values()
            if row.get("material_group")
        }),
        "failed_count": len(failed_rows),
        "no_criteria_count": no_criteria_count,
        "report_file": report_path.name,
        "analysis_file": analysis_path.name,
    }
    uploads.insert(0, upload_record)

    progress(94, "Saving analyzed data.")
    write_json(MATERIALS_FILE, materials)
    write_json(UPLOADS_FILE, uploads)

    progress(98, "Refreshing dashboard summary.")
    return {
        "upload": upload_record,
        "failed_rows": failed_rows,
        "summary": build_summary(),
    }


INVENTORY_BUCKET_COLUMNS = {
    "qty_0_30": ["Qty in Days( 0 - 30 )", "Qty 0-30", "0-30 Days Qty"],
    "qty_31_60": ["Qty in Days( 31 - 60 )", "Qty 31-60", "31-60 Days Qty"],
    "qty_61_90": ["Qty in Days( 61 - 90 )", "Qty 61-90", "61-90 Days Qty"],
    "qty_91_180": ["Qty in Days( 91 - 180 )", "Qty 91-180", "91-180 Days Qty"],
    "qty_181_365": ["Qty in Days( 181 - 365 )", "Qty 181-365", "181-365 Days Qty"],
    "qty_1_2_years": ["Qty ( 1 - 2 Yr )", "Qty 1-2 Yr", "1-2 Years Qty"],
    "qty_over_2_years": ["Qty ( > 2 Yrs )", "Qty >2 Yr", "Qty > 2 Years", ">2 Years Qty"],
}


def inventory_analysis_for_upload(upload):
    if not upload:
        return {}
    file_name = upload.get("analysis_file", "")
    path = INVENTORY_ANALYSIS_DIR / file_name
    return read_json(path) if file_name and path.exists() else {}


def latest_inventory_analysis():
    uploads = read_json(INVENTORY_UPLOADS_FILE)
    upload = uploads[0] if uploads else None
    return upload, inventory_analysis_for_upload(upload)


def process_inventory_upload(file_path, original_name, uploaded_by="", progress_callback=None):
    def progress(percent, message):
        if progress_callback:
            progress_callback(percent, message)

    progress(8, "Opening inventory workbook.")
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Invalid Excel file: {exc}")
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ValueError("Inventory Excel sheet is empty.")
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    indexes = {
        "plant": find_column(headers, ["Plant", "Plant Code"]),
        "material": find_column(headers, ["Part Number", "Material", "Material Code", "Part No"]),
        "description": find_column(headers, ["Description", "Material Description", "Part Description"]),
        "total_stock": find_column(headers, ["Total Stock", "Current Stock", "Stock Qty", "Stock Quantity"]),
        "total_value": find_column(headers, ["Total Value", "Stock Value", "Valuated Stock Value"]),
    }
    indexes.update({key: find_column(headers, candidates) for key, candidates in INVENTORY_BUCKET_COLUMNS.items()})
    required = [label for label in ("plant", "material", "total_stock") if indexes[label] is None]
    if required:
        raise ValueError(
            "Missing required inventory columns: "
            + ", ".join(required)
            + ". Expected columns include Plant, Part Number/Material, and Total Stock."
        )

    progress(15, "Reading inventory age buckets.")
    total_rows = max((ws.max_row or 1) - 1, 1)
    plants = {}
    parts = {}
    for row_index, row in enumerate(rows, start=2):
        processed = row_index - 1
        if processed == 1 or processed % 100 == 0 or processed >= total_rows:
            progress(15 + min(70, int(processed * 70 / total_rows)), f"Analyzed {min(processed, total_rows)} of {total_rows} inventory rows.")
        if not row or all(value is None for value in row):
            continue

        def cell(key):
            index = indexes.get(key)
            return row[index] if index is not None and index < len(row) else None

        plant = str(cell("plant") or "Unassigned").strip()
        material = str(cell("material") or "").strip()
        if not material:
            continue
        description = str(cell("description") or "").strip()
        stock = parse_number(cell("total_stock"))
        value = parse_number(cell("total_value"))
        bucket_values = {key: parse_number(cell(key)) for key in INVENTORY_BUCKET_COLUMNS}
        over_six_months = bucket_values["qty_181_365"] + bucket_values["qty_1_2_years"] + bucket_values["qty_over_2_years"]
        plant_summary = plants.setdefault(plant, {
            "plant": plant,
            "parts_count": 0,
            "total_stock": 0,
            "total_value": 0,
            "qty_0_30": 0,
            "qty_0_90": 0,
            "qty_31_60": 0,
            "qty_61_90": 0,
            "qty_91_180": 0,
            "qty_181_365": 0,
            "qty_1_2_years": 0,
            "qty_over_2_years": 0,
            "over_six_months": 0,
            "attention_parts": 0,
        })
        plant_summary["total_stock"] += stock
        plant_summary["total_value"] += value
        plant_summary["qty_0_90"] += (
            bucket_values["qty_0_30"]
            + bucket_values["qty_31_60"]
            + bucket_values["qty_61_90"]
        )
        plant_summary["over_six_months"] += over_six_months
        for key, bucket_value in bucket_values.items():
            plant_summary[key] += bucket_value
        part_key = f"{plant}::{material}"
        part = parts.setdefault(part_key, {
            "plant": plant,
            "material": material,
            "description": description,
            "total_stock": 0,
            "total_value": 0,
            "qty_0_90": 0,
            "qty_over_2_years": 0,
            "over_six_months": 0,
        })
        part["total_stock"] += stock
        part["total_value"] += value
        part["qty_0_90"] += (
            bucket_values["qty_0_30"]
            + bucket_values["qty_31_60"]
            + bucket_values["qty_61_90"]
        )
        part["qty_over_2_years"] += bucket_values["qty_over_2_years"]
        part["over_six_months"] += over_six_months

    part_rows = list(parts.values())
    for plant_summary in plants.values():
        plant_summary["parts_count"] = sum(1 for row in part_rows if row["plant"] == plant_summary["plant"])
        plant_summary["attention_parts"] = sum(
            1 for row in part_rows
            if row["plant"] == plant_summary["plant"] and row["qty_over_2_years"] > 0
        )
        for key, value in list(plant_summary.items()):
            if isinstance(value, float):
                plant_summary[key] = round(value, 2)

    progress(89, "Preparing inventory trend and part analysis.")
    upload_id = uuid.uuid4().hex[:10]
    part_rows.sort(key=lambda row: (-row["qty_over_2_years"], -row["over_six_months"], row["material"]))
    analysis = {
        "upload_id": upload_id,
        "source_file": original_name,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "plants": sorted(plants.values(), key=lambda row: row["plant"]),
        "parts": part_rows,
        "method": "Age buckets are calculated directly from the inventory Excel. Consumption forecasting requires issue or consumption history.",
    }
    analysis_path = INVENTORY_ANALYSIS_DIR / f"inventory_analysis_{upload_id}.json"
    write_json(analysis_path, analysis)
    uploads = [
        item for item in read_json(INVENTORY_UPLOADS_FILE)
        if item.get("file_name") != original_name
    ]
    upload = {
        "id": upload_id,
        "file_name": original_name,
        "uploaded_by": uploaded_by,
        "uploaded_at": analysis["generated_at"],
        "plants_count": len(plants),
        "parts_count": len(part_rows),
        "analysis_file": analysis_path.name,
    }
    uploads.insert(0, upload)
    write_json(INVENTORY_UPLOADS_FILE, uploads[:24])
    progress(98, "Inventory analysis ready.")
    return {"upload": upload, "analysis": analysis}


def inventory_dashboard_payload():
    upload, analysis = latest_inventory_analysis()
    uploads = read_json(INVENTORY_UPLOADS_FILE)
    history = []
    for item in reversed(uploads[:12]):
        snapshot = inventory_analysis_for_upload(item)
        plants = snapshot.get("plants", [])
        history.append({
            "uploaded_at": item.get("uploaded_at", ""),
            "file_name": item.get("file_name", ""),
            "total_value": round(sum(parse_number(row.get("total_value")) for row in plants), 2),
            "over_six_months": round(sum(parse_number(row.get("over_six_months")) for row in plants), 2),
            "over_two_years": round(sum(parse_number(row.get("qty_over_2_years")) for row in plants), 2),
        })
    return {
        "upload": upload,
        "plants": analysis.get("plants", []),
        "parts": analysis.get("parts", [])[:150],
        "history": history,
        "method": analysis.get("method", "Upload an agewise inventory Excel to start analysis."),
    }


def build_summary():
    criteria = read_json(CRITERIA_FILE)
    uploads = read_json(UPLOADS_FILE)
    group_data = get_material_group_data()
    master_materials_count = len(group_data.get("materials", {}))
    low_stock = 0
    category_cards = {}
    for category, label in CATEGORIES.items():
        stock_zone_analysis = stock_zone_analysis_for_category(category)
        stock_zone_counts = stock_zone_analysis.get("counts", {})
        stock_zone_groups = stock_zone_analysis.get("material_groups", {})
        stock_zone_dashboard = {
            "total": stock_zone_analysis.get("total_materials", 0),
            "uploaded": stock_zone_analysis.get("total_materials", 0),
            "criteria": stock_zone_analysis.get("eligible_materials", 0),
            "critical": stock_zone_analysis.get("critical_materials", 0),
            "active": stock_zone_analysis.get("stock_parts_count", stock_zone_analysis.get("eligible_materials", 0)),
            "red": stock_zone_counts.get("red", 0),
            "yellow": stock_zone_counts.get("yellow", 0),
            "green": stock_zone_counts.get("green", 0),
            "not_eligible": stock_zone_counts.get("not_eligible", 0),
        }
        category_materials = get_category_materials(category)
        upload_summary = upload_analysis_summary(category)
        critical_defaults = category_critical_defaults(category)
        category_criteria = [
            item for item in criteria.values()
            if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
        ]
        default_criteria_count = len([
            item for item in category_materials
            if str(item.get("material", "")) in critical_defaults
        ])
        category_uploads = [
            item for item in uploads
            if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
        ]
        latest_upload = category_uploads[0] if category_uploads else None
        card_low_stock = 0
        group_cards = {
            code: {
                "code": code,
                "label": group.get("label", code),
                "source_parts_count": group.get("source_parts_count", 0),
                "materials_count": 0,
                "criteria_count": 0,
                "low_stock_count": 0,
                "active_stocks_count": 0,
                "zone_analysis": stock_zone_groups.get(code, {}),
            }
            for code, group in group_data.get("groups", {}).items()
        }
        for material in category_materials:
            code = str(material.get("material", ""))
            group_code = material.get("material_group") or material_group_for(code)
            if group_code:
                group_cards.setdefault(group_code, {
                    "code": group_code,
                    "label": group_code,
                    "source_parts_count": 0,
                    "materials_count": 0,
                    "criteria_count": 0,
                    "low_stock_count": 0,
                    "active_stocks_count": 0,
                    "zone_analysis": stock_zone_groups.get(group_code, {}),
                })
                group_cards[group_code]["materials_count"] += 1
                if parse_number(material.get("last_quantity")) < 0:
                    group_cards[group_code]["active_stocks_count"] += 1
            settings = get_material_criteria(criteria, category, code)
            if not settings or not settings.get("active", True):
                continue
            if group_code:
                group_cards[group_code]["criteria_count"] += 1
            if parse_number(material.get("current_stock")) <= parse_number(settings.get("minimum_stock")):
                card_low_stock += 1
                if group_code:
                    group_cards[group_code]["low_stock_count"] += 1
        if latest_upload:
            card_low_stock = upload_summary["critical"]
            for group_code, counts in upload_summary["groups"].items():
                group_cards.setdefault(group_code, {
                    "code": group_code,
                    "label": group_code,
                    "source_parts_count": 0,
                    "materials_count": 0,
                    "criteria_count": 0,
                    "low_stock_count": 0,
                    "active_stocks_count": 0,
                    "zone_analysis": stock_zone_groups.get(group_code, {}),
                })
                if category not in STOCK_ZONE_CATEGORIES:
                    group_cards[group_code]["materials_count"] = counts["uploaded"]
                    group_cards[group_code]["active_stocks_count"] = counts["active"]
                    group_cards[group_code]["low_stock_count"] = counts["critical"]
                    group_cards[group_code]["zone_analysis"] = {
                        **group_cards[group_code].get("zone_analysis", {}),
                        "uploaded": counts["uploaded"],
                        "active": counts["active"],
                        "critical": counts["critical"],
                    }
            if category not in STOCK_ZONE_CATEGORIES:
                stock_zone_dashboard = {
                    **stock_zone_dashboard,
                    "uploaded": upload_summary["uploaded"],
                    "active": upload_summary["active"],
                    "critical": upload_summary["critical"],
                    "red": upload_summary["critical"],
                    "yellow": 0,
                    "green": max(0, upload_summary["active"] - upload_summary["critical"]),
                }
                for group_code, counts in upload_summary["groups"].items():
                    group_cards[group_code]["zone_analysis"] = {
                        **group_cards[group_code].get("zone_analysis", {}),
                        "red": counts["critical"],
                        "yellow": 0,
                        "green": max(0, counts["active"] - counts["critical"]),
                        "active": counts["active"],
                        "critical": counts["critical"],
                    }
            else:
                card_low_stock = stock_zone_dashboard["critical"]
        if category in STOCK_ZONE_CATEGORIES:
            card_low_stock = stock_zone_dashboard["critical"]
        category_cards[category] = {
            "key": category,
            "label": label,
            "materials_count": upload_summary["uploaded"] if latest_upload else len(category_materials),
            "criteria_count": len(category_criteria) + default_criteria_count,
            "master_parts_count": master_materials_count if category == DEFAULT_CATEGORY else stock_zone_analysis.get("total_materials", 0),
            "low_stock_count": card_low_stock,
            "uploads_count": len(category_uploads),
            "latest_upload": latest_upload,
            "active_stocks_count": stock_zone_dashboard["active"] if category in STOCK_ZONE_CATEGORIES else (upload_summary["active"] if latest_upload else 0),
            "active_quantity_used": latest_upload.get("used_quantity_total", 0) if latest_upload else 0,
            "zone_counts": stock_zone_analysis.get("counts", {"red": 0, "yellow": 0, "green": 0}),
            "zone_analysis": stock_zone_dashboard,
            "zone_total": stock_zone_analysis.get("total_materials", 0),
            "material_groups": group_cards,
        }
    low_stock = sum(card["low_stock_count"] for card in category_cards.values())
    return {
        "materials_count": sum(len(get_category_materials(category)) for category in CATEGORIES),
        "criteria_count": len(criteria),
        "uploads_count": len(uploads),
        "low_stock_count": low_stock,
        "latest_upload": uploads[0] if uploads else None,
        "categories": category_cards,
    }


def search_materials(query, category=DEFAULT_CATEGORY, material_group=""):
    category = normalize_category(category)
    criteria = read_json(CRITERIA_FILE)
    normalized = query.strip().lower()
    results = []
    for material in get_category_materials(category):
        code = str(material.get("material", ""))
        group_code = material.get("material_group") or material_group_for(code)
        if material_group and group_code != material_group:
            continue
        description = str(material.get("description", ""))
        haystack = f"{code} {description}".lower()
        if normalized and normalized not in haystack:
            continue
        setting = get_material_criteria(criteria, category, code) or {}
        critical_default = default_critical_for(code, category)
        results.append({
            "material": code,
            "description": description,
            "category": category,
            "material_group": group_code,
            "current_stock": material.get("current_stock", 0),
            "critical_value": setting.get("minimum_stock", critical_default.get("critical_value", "")),
            "net_consumption": setting.get("net_consumption", critical_default.get("net_consumption", "")),
            "last_entry_date": material.get("last_entry_date", ""),
            "minimum_stock": setting.get("minimum_stock", ""),
            "reorder_quantity": setting.get("reorder_quantity", ""),
            "has_criteria": bool(setting),
            "keep_stock": bool(setting.get("keep_stock")),
        })
    results.sort(key=lambda item: (not item["has_criteria"], item["material"]))
    return results[:25]


class AppHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/api/me":
            return json_response(self, 200, {"user": current_user(self)})
        if path == "/api/summary":
            if not require_user(self):
                return
            return json_response(self, 200, build_summary())
        if path == "/api/criteria":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            category = normalize_category(query.get("category", [DEFAULT_CATEGORY])[0])
            material_group = query.get("material_group", [""])[0]
            criteria = read_json(CRITERIA_FILE)
            filtered = {
                key: value for key, value in criteria.items()
                if normalize_category(value.get("category", DEFAULT_CATEGORY)) == category
            }
            filtered = {
                key: {**value, "material_group": material_group_for(value.get("material"))}
                for key, value in filtered.items()
            }
            if material_group:
                filtered = {
                    key: value for key, value in filtered.items()
                    if material_group_for(value.get("material")) == material_group
                }
            return json_response(self, 200, filtered)
        if path == "/api/materials":
            if not require_user(self):
                return
            parsed_query = parse_qs(parsed.query)
            query = parsed_query.get("q", [""])[0]
            category = normalize_category(parsed_query.get("category", [DEFAULT_CATEGORY])[0])
            material_group = parsed_query.get("material_group", [""])[0]
            if query:
                return json_response(self, 200, search_materials(query, category, material_group))
            rows = [
                item for item in get_category_materials(category)
                if not material_group or (item.get("material_group") or material_group_for(item.get("material"))) == material_group
            ]
            return json_response(self, 200, rows[:100])
        if path == "/api/material-groups":
            if not require_user(self):
                return
            group_data = get_material_group_data()
            groups = sorted(group_data.get("groups", {}).values(), key=lambda item: item.get("code", ""))
            return json_response(self, 200, {"groups": groups})
        if path == "/api/group-critical":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            category = normalize_category(query.get("category", [DEFAULT_CATEGORY])[0])
            material_group = query.get("material_group", [""])[0]
            critical_stock = query.get("critical_stock", [""])[0]
            return json_response(self, 200, analyze_group_critical(category, material_group, critical_stock))
        if path == "/api/zone-parts":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            category = normalize_category(query.get("category", [DEFAULT_CATEGORY])[0])
            zone = query.get("zone", ["red"])[0]
            material_group = query.get("material_group", [""])[0]
            rows = zone_part_rows(category, zone, material_group)
            return json_response(self, 200, {"zone": normalize_zone(zone), "value": len(rows), "rows": rows})
        if path == "/api/zone-export":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            category = normalize_category(query.get("category", [DEFAULT_CATEGORY])[0])
            zone = normalize_zone(query.get("zone", ["red"])[0]) or "red"
            material_group = query.get("material_group", [""])[0]
            wb = build_zone_export(category, zone, material_group)
            output = io.BytesIO()
            wb.save(output)
            body = output.getvalue()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{category}_{zone}_zone_parts.xlsx"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if path == "/api/stock-overrides":
            if not require_user(self):
                return
            return json_response(self, 200, read_json(STOCK_OVERRIDES_FILE))
        if path == "/api/uploads":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            category = query.get("category", [""])[0]
            uploads = read_json(UPLOADS_FILE)
            if category:
                category = normalize_category(category)
                uploads = [
                    item for item in uploads
                    if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
                ]
            return json_response(self, 200, uploads)
        if path == "/api/inventory-analysis":
            if not require_user(self):
                return
            return json_response(self, 200, inventory_dashboard_payload())
        if path == "/api/inventory-uploads":
            if not require_user(self):
                return
            return json_response(self, 200, read_json(INVENTORY_UPLOADS_FILE))
        if path == "/api/upload-progress":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            job_id = query.get("job_id", [""])[0]
            return json_response(self, 200, get_upload_progress(job_id))
        if path == "/api/analysis":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            upload_id = query.get("upload_id", [""])[0]
            category = query.get("category", [""])[0]
            material_group = query.get("material_group", [""])[0]
            uploads = read_json(UPLOADS_FILE)
            if category:
                category = normalize_category(category)
                uploads = [
                    item for item in uploads
                    if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
                ]
            upload = next((item for item in uploads if item.get("id") == upload_id), uploads[0] if uploads else None)
            if not upload:
                return json_response(self, 200, {"upload": None, "failed_rows": [], "no_criteria_count": 0})
            analysis_file = upload.get("analysis_file") or f"analysis_{upload['id']}.json"
            analysis_path = ANALYSIS_DIR / analysis_file
            if not analysis_path.exists():
                return json_response(self, 200, {"upload": upload, "failed_rows": [], "no_criteria_count": 0})
            analysis = read_json(analysis_path)
            failed_rows = analysis.get("failed_rows", [])
            if material_group:
                failed_rows = [
                    row for row in failed_rows
                    if material_group_for(row.get("material")) == material_group
                ]
            failed_rows = [
                {**row, "material_group": row.get("material_group") or material_group_for(row.get("material"))}
                for row in failed_rows
            ]
            return json_response(self, 200, {
                "upload": upload,
                "failed_rows": failed_rows,
                "no_criteria_count": analysis.get("no_criteria_count", 0),
                "no_criteria_rows": analysis.get("no_criteria_rows", []),
                "generated_at": analysis.get("generated_at"),
            })
        if path == "/api/no-criteria-export":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            upload_id = query.get("upload_id", [""])[0]
            category = query.get("category", [""])[0]
            material_group = query.get("material_group", [""])[0]
            uploads = read_json(UPLOADS_FILE)
            if category:
                category = normalize_category(category)
                uploads = [
                    item for item in uploads
                    if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
                ]
            upload = next((item for item in uploads if item.get("id") == upload_id), uploads[0] if uploads else None)
            if not upload:
                return json_response(self, 404, {"error": "Upload not found."})
            analysis = analysis_for_upload(upload)
            rows = analysis.get("no_criteria_rows", [])
            if material_group:
                rows = [
                    row for row in rows
                    if (row.get("material_group") or material_group_for(row.get("material"))) == material_group
                ]
            rows = [
                {**row, "material_group": row.get("material_group") or material_group_for(row.get("material"))}
                for row in rows
            ]
            wb = build_no_criteria_export(rows)
            output = io.BytesIO()
            wb.save(output)
            body = output.getvalue()
            safe_category = normalize_category(upload.get("category", DEFAULT_CATEGORY))
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{safe_category}_no_criteria_items_{upload.get("id", "latest")}.xlsx"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if path == "/api/download":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            name = query.get("file", [""])[0]
            report_path = (REPORT_DIR / name).resolve()
            if REPORT_DIR.resolve() not in report_path.parents or not report_path.exists():
                return json_response(self, 404, {"error": "Report not found."})
            body = report_path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{report_path.name}"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        return self.serve_static(path)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/signup":
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            name = str(payload.get("name", "")).strip()
            email = str(payload.get("email", "")).strip().lower()
            password = str(payload.get("password", ""))
            role = str(payload.get("role", "employee")).strip().lower()
            allowed_roles = {"admin", "employee", "uploader", "criteria"}
            if role not in allowed_roles:
                return json_response(self, 400, {"error": "Invalid role."})
            try:
                user = create_user(name, email, password, role)
            except ValueError as exc:
                return json_response(self, 409, {"error": str(exc)})
            session_id = create_session(email)
            body = json.dumps({"user": user}).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Set-Cookie", f"session_id={session_id}; Path=/; SameSite=Lax")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/login":
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            email = str(payload.get("email", "")).strip().lower()
            password = str(payload.get("password", ""))
            user = authenticate_user(email, password)
            if not user:
                return json_response(self, 401, {"error": "Invalid email or password."})
            session_id = create_session(email)
            body = json.dumps({"user": user}).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Set-Cookie", f"session_id={session_id}; Path=/; SameSite=Lax")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/logout":
            cookies = parse_cookies(self.headers.get("Cookie"))
            session_id = cookies.get("session_id")
            delete_session(session_id)
            body = json.dumps({"ok": True}).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Set-Cookie", "session_id=; Path=/; Max-Age=0; SameSite=Lax")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/criteria":
            if not require_role(self, "admin"):
                return
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            material = str(payload.get("material", "")).strip()
            category = normalize_category(payload.get("category", DEFAULT_CATEGORY))
            if not material:
                return json_response(self, 400, {"error": "Material code is required."})
            criteria = read_json(CRITERIA_FILE)
            key = criteria_key(category, material)
            for existing_key, existing_value in list(criteria.items()):
                if existing_key != key and criteria_matches(existing_value, category, material):
                    del criteria[existing_key]
            criteria[key] = {
                "material": material,
                "description": str(payload.get("description", "")).strip(),
                "category": category,
                "minimum_stock": parse_number(payload.get("minimum_stock")),
                "reorder_quantity": parse_number(payload.get("reorder_quantity")),
                "active": bool(payload.get("active", True)),
                "keep_stock": bool(payload.get("keep_stock")),
            }
            write_json(CRITERIA_FILE, criteria)
            global STOCK_ZONE_ANALYSIS_CACHE
            STOCK_ZONE_ANALYSIS_CACHE = None
            sync_stock_zone_analysis()
            return json_response(self, 200, {"criteria": criteria[key], "summary": build_summary()})

        if parsed.path == "/api/group-critical":
            if not require_role(self, "admin"):
                return
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            category = normalize_category(payload.get("category", DEFAULT_CATEGORY))
            material_group = str(payload.get("material_group", "")).strip()
            if not material_group:
                return json_response(self, 400, {"error": "Please select a material group first."})
            group_settings = read_json(GROUP_CRITERIA_FILE)
            key = group_criteria_key(category, material_group)
            group_settings[key] = {
                "category": category,
                "material_group": material_group,
                "critical_stock": parse_number(payload.get("critical_stock")),
                "reorder_quantity": parse_number(payload.get("reorder_quantity")),
                "active": bool(payload.get("active", True)),
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            write_json(GROUP_CRITERIA_FILE, group_settings)
            analysis = analyze_group_critical(category, material_group, group_settings[key]["critical_stock"])
            return json_response(self, 200, {"criteria": group_settings[key], "analysis": analysis})

        if parsed.path == "/api/stock-overrides":
            if not require_role(self, "admin"):
                return
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            material = str(payload.get("material", "")).strip()
            action = "remove" if payload.get("action") == "remove" else "keep"
            category = normalize_category(payload.get("category", DEFAULT_CATEGORY))
            if not material:
                return json_response(self, 400, {"error": "Material code is required."})
            overrides = read_json(STOCK_OVERRIDES_FILE)
            overrides.setdefault("keep", {})
            overrides.setdefault("remove", {})
            overrides["keep"].pop(material, None)
            overrides["remove"].pop(material, None)
            overrides[action][material] = {
                "material": material,
                "description": str(payload.get("description", "")).strip(),
                "category": category,
                "material_group": str(payload.get("material_group", "")).strip(),
                "action": action,
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            write_json(STOCK_OVERRIDES_FILE, overrides)
            refresh_stock_after_override()
            return json_response(self, 200, {"overrides": overrides, "summary": build_summary()})

        if parsed.path == "/api/upload":
            if not require_user(self):
                return
            form = parse_multipart_form(self)
            file_item = form["files"].get("file")
            if not file_item or not file_item.get("filename"):
                return json_response(self, 400, {"error": "Please choose an Excel file."})
            category = normalize_category(form["fields"].get("category", DEFAULT_CATEGORY))
            job_id = str(form["fields"].get("job_id", "")).strip() or uuid.uuid4().hex
            upload_name = Path(file_item["filename"]).name
            saved_path = UPLOAD_DIR / f"{uuid.uuid4().hex}_{upload_name}"
            set_upload_progress(job_id, 3, "Saving uploaded Excel file.", category=category)
            with saved_path.open("wb") as output:
                output.write(file_item["content"])
            try:
                user = current_user(self) or {}
                result = process_upload(
                    saved_path,
                    upload_name,
                    category,
                    user.get("email", ""),
                    progress_callback=lambda percent, message: set_upload_progress(job_id, percent, message, category=category),
                )
                set_upload_progress(job_id, 100, "Excel processing completed.", state="done", category=category)
                return json_response(self, 200, result)
            except Exception as exc:
                set_upload_progress(job_id, 100, str(exc), state="error", category=category)
                return json_response(self, 400, {"error": str(exc)})

        if parsed.path == "/api/inventory-upload":
            if not require_user(self):
                return
            form = parse_multipart_form(self)
            file_item = form["files"].get("file")
            if not file_item or not file_item.get("filename"):
                return json_response(self, 400, {"error": "Please choose an inventory Excel file."})
            job_id = str(form["fields"].get("job_id", "")).strip() or uuid.uuid4().hex
            upload_name = Path(file_item["filename"]).name
            saved_path = UPLOAD_DIR / f"inventory_{uuid.uuid4().hex}_{upload_name}"
            set_upload_progress(job_id, 3, "Saving inventory Excel file.", category="inventory")
            with saved_path.open("wb") as output:
                output.write(file_item["content"])
            try:
                user = current_user(self) or {}
                result = process_inventory_upload(
                    saved_path,
                    upload_name,
                    user.get("email", ""),
                    progress_callback=lambda percent, message: set_upload_progress(job_id, percent, message, category="inventory"),
                )
                set_upload_progress(job_id, 100, "Inventory analysis completed.", state="done", category="inventory")
                return json_response(self, 200, result)
            except Exception as exc:
                set_upload_progress(job_id, 100, str(exc), state="error", category="inventory")
                return json_response(self, 400, {"error": str(exc)})

        return json_response(self, 404, {"error": "Unknown endpoint."})

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/criteria":
            if not require_role(self, "admin"):
                return
            query = parse_qs(parsed.query)
            category = normalize_category(query.get("category", [DEFAULT_CATEGORY])[0])
            material = str(query.get("material", [""])[0]).strip()
            if not material:
                return json_response(self, 400, {"error": "Material code is required."})
            criteria = read_json(CRITERIA_FILE)
            deleted = False
            key = criteria_key(category, material)
            if key in criteria:
                del criteria[key]
                deleted = True
            for existing_key, existing_value in list(criteria.items()):
                if criteria_matches(existing_value, category, material):
                    del criteria[existing_key]
                    deleted = True
            if not deleted:
                return json_response(self, 404, {"error": "Criteria not found."})
            write_json(CRITERIA_FILE, criteria)
            global STOCK_ZONE_ANALYSIS_CACHE
            STOCK_ZONE_ANALYSIS_CACHE = None
            sync_stock_zone_analysis()
            return json_response(self, 200, {"summary": build_summary()})
        if parsed.path == "/api/stock-overrides":
            if not require_role(self, "admin"):
                return
            query = parse_qs(parsed.query)
            material = str(query.get("material", [""])[0]).strip()
            if not material:
                return json_response(self, 400, {"error": "Material code is required."})
            overrides = read_json(STOCK_OVERRIDES_FILE)
            overrides.setdefault("keep", {}).pop(material, None)
            overrides.setdefault("remove", {}).pop(material, None)
            write_json(STOCK_OVERRIDES_FILE, overrides)
            refresh_stock_after_override()
            return json_response(self, 200, {"overrides": overrides, "summary": build_summary()})
        return json_response(self, 404, {"error": "Unknown endpoint."})

    def serve_static(self, path):
        if path == "/":
            path = "/index.html"
        file_path = (STATIC_DIR / path.lstrip("/")).resolve()
        if STATIC_DIR.resolve() not in file_path.parents and file_path != STATIC_DIR.resolve():
            return json_response(self, 403, {"error": "Forbidden."})
        if not file_path.exists() or not file_path.is_file():
            return json_response(self, 404, {"error": "Not found."})
        body = file_path.read_bytes()
        content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        sys.stderr.write("%s - %s\n" % (self.address_string(), format % args))


if __name__ == "__main__":
    ensure_dirs()
    server = ThreadingHTTPServer(("127.0.0.1", 8000), AppHandler)
    print("Tata Inventory Criteria Checker running at http://127.0.0.1:8000")
    server.serve_forever()
