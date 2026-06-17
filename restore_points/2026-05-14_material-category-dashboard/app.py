from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
import cgi
import json
import mimetypes
import shutil
import sys
import uuid
from datetime import datetime

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
REPORT_DIR = DATA_DIR / "reports"
ANALYSIS_DIR = DATA_DIR / "analysis"
CRITERIA_FILE = DATA_DIR / "criteria.json"
GROUP_CRITERIA_FILE = DATA_DIR / "group_criteria.json"
UPLOADS_FILE = DATA_DIR / "uploads.json"
MATERIALS_FILE = DATA_DIR / "materials.json"
MATERIAL_GROUPS_FILE = DATA_DIR / "material_groups.json"
MATERIAL_GROUP_SOURCE = BASE_DIR.parent / "updated" / "Material group 3002 and 3004.xlsx"

REQUIRED_COLUMNS = [
    "Purchase Order Date",
    "Entry Date",
    "Material",
    "Material Description",
    "Quantity",
    "Days Between",
    "Valuated Stock",
]

DEMO_USERS = {
    "admin@tatamotors.com": {
        "name": "Admin",
        "email": "admin@tatamotors.com",
        "password": "admin123",
        "role": "admin",
    },
    "employee@tatamotors.com": {
        "name": "Employee",
        "email": "employee@tatamotors.com",
        "password": "employee123",
        "role": "employee",
    },
    "uploader@tatamotors.com": {
        "name": "Tata Uploader",
        "email": "uploader@tatamotors.com",
        "password": "upload123",
        "role": "employee",
    },
    "criteria@tatamotors.com": {
        "name": "Criteria Manager",
        "email": "criteria@tatamotors.com",
        "password": "criteria123",
        "role": "admin",
    },
}

SESSIONS = {}
MATERIAL_GROUP_CACHE = None

CATEGORIES = {
    "ml_spare": "M/L Spare",
    "tools": "Tools",
}
DEFAULT_CATEGORY = "ml_spare"


def ensure_dirs():
    for path in [STATIC_DIR, DATA_DIR, UPLOAD_DIR, REPORT_DIR, ANALYSIS_DIR]:
        path.mkdir(parents=True, exist_ok=True)
    for file_path, empty in [
        (CRITERIA_FILE, {}),
        (GROUP_CRITERIA_FILE, {}),
        (UPLOADS_FILE, []),
        (MATERIALS_FILE, {}),
        (MATERIAL_GROUPS_FILE, {"groups": {}, "materials": {}}),
    ]:
        if not file_path.exists():
            file_path.write_text(json.dumps(empty, indent=2), encoding="utf-8")
    sync_material_groups()


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path, payload):
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def sync_material_groups():
    global MATERIAL_GROUP_CACHE
    if not MATERIAL_GROUP_SOURCE.exists():
        return
    source_mtime = MATERIAL_GROUP_SOURCE.stat().st_mtime
    existing = read_json(MATERIAL_GROUPS_FILE) if MATERIAL_GROUPS_FILE.exists() else {}
    if existing.get("source_mtime") == source_mtime:
        MATERIAL_GROUP_CACHE = existing
        return

    wb = load_workbook(MATERIAL_GROUP_SOURCE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    try:
        material_idx = headers.index("Material")
        group_idx = headers.index("Material Group")
    except ValueError:
        return
    description_idx = headers.index("Short Text") if "Short Text" in headers else None

    groups = {}
    material_lookup = {}
    for row in rows:
        material = str(row[material_idx] or "").strip()
        material_group = str(row[group_idx] or "").strip()
        if not material or not material_group:
            continue
        description = str(row[description_idx] or "").strip() if description_idx is not None else ""
        groups.setdefault(material_group, {
            "code": material_group,
            "label": material_group,
            "source_parts_count": 0,
        })
        groups[material_group]["source_parts_count"] += 1
        material_lookup[material] = {
            "material": material,
            "material_group": material_group,
            "description": description,
        }

    payload = {
        "source_file": str(MATERIAL_GROUP_SOURCE),
        "source_mtime": source_mtime,
        "groups": dict(sorted(groups.items())),
        "materials": material_lookup,
    }
    write_json(MATERIAL_GROUPS_FILE, payload)
    MATERIAL_GROUP_CACHE = payload


def get_material_group_data():
    global MATERIAL_GROUP_CACHE
    if MATERIAL_GROUP_CACHE is None:
        sync_material_groups()
    if MATERIAL_GROUP_CACHE is None:
        MATERIAL_GROUP_CACHE = read_json(MATERIAL_GROUPS_FILE)
    return MATERIAL_GROUP_CACHE


def material_group_for(material):
    material_data = get_material_group_data().get("materials", {}).get(str(material), {})
    return material_data.get("material_group", "")


def parse_number(value, default=0):
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def normalize_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def normalize_category(value):
    value = str(value or "").strip().lower()
    return value if value in CATEGORIES else DEFAULT_CATEGORY


def criteria_key(category, material):
    return f"{normalize_category(category)}::{material}"


def group_criteria_key(category, material_group):
    return f"{normalize_category(category)}::{material_group}"


def criteria_matches(setting, category, material):
    setting_category = normalize_category(setting.get("category", DEFAULT_CATEGORY))
    return setting.get("material") == material and setting_category == normalize_category(category)


def get_material_criteria(criteria, category, material):
    category = normalize_category(category)
    direct = criteria.get(criteria_key(category, material))
    if direct:
        return direct
    legacy = criteria.get(material)
    if legacy and normalize_category(legacy.get("category", DEFAULT_CATEGORY)) == category:
        return legacy
    for setting in criteria.values():
        if criteria_matches(setting, category, material):
            return setting
    return None


def material_description_for(material):
    group_data = get_material_group_data().get("materials", {}).get(str(material), {})
    return group_data.get("description", "")


def analyze_group_critical(category, material_group, critical_stock=None):
    category = normalize_category(category)
    materials = read_json(MATERIALS_FILE)
    group_settings = read_json(GROUP_CRITERIA_FILE)
    saved = group_settings.get(group_criteria_key(category, material_group), {})
    if critical_stock is None or critical_stock == "":
        critical_stock = saved.get("critical_stock", "")
    critical_stock = parse_number(critical_stock, None)

    group_rows = []
    critical_rows = []
    for material in materials.values():
        if normalize_category(material.get("category", DEFAULT_CATEGORY)) != category:
            continue
        code = str(material.get("material", ""))
        group_code = material.get("material_group") or material_group_for(code)
        if material_group and group_code != material_group:
            continue
        current_stock = parse_number(material.get("current_stock"))
        description = material.get("description") or material_description_for(code)
        row = {
            "material": code,
            "material_group": group_code,
            "description": description,
            "current_stock": current_stock,
            "last_quantity": material.get("last_quantity", 0),
            "movement_type": material.get("movement_type", ""),
            "last_entry_date": material.get("last_entry_date", ""),
        }
        group_rows.append(row)
        if critical_stock is not None and current_stock <= critical_stock:
            critical_rows.append({
                **row,
                "reason": f"Current stock {current_stock:g} is at or below critical value {critical_stock:g}",
            })

    stocks = sorted(parse_number(row.get("current_stock")) for row in group_rows)
    examples = []
    if stocks:
        examples = sorted(set([
            stocks[0],
            stocks[len(stocks) // 4],
            stocks[len(stocks) // 2],
            5,
        ]))
    critical_rows.sort(key=lambda item: (item["current_stock"], item["material"]))
    return {
        "category": category,
        "material_group": material_group,
        "saved_criteria": saved,
        "critical_stock": critical_stock,
        "total_parts": len(group_rows),
        "critical_count": len(critical_rows),
        "examples": examples[:4],
        "rows": critical_rows[:100],
    }


def json_response(handler, status, payload):
    body = json.dumps(payload, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def parse_cookies(header):
    cookies = {}
    if not header:
        return cookies
    for part in header.split(";"):
        if "=" not in part:
            continue
        key, value = part.strip().split("=", 1)
        cookies[key] = value
    return cookies


def current_user(handler):
    cookies = parse_cookies(handler.headers.get("Cookie"))
    session_id = cookies.get("session_id")
    if not session_id:
        return None
    email = SESSIONS.get(session_id)
    if not email:
        return None
    user = DEMO_USERS.get(email)
    if not user:
        return None
    return {key: value for key, value in user.items() if key != "password"}


def require_user(handler):
    user = current_user(handler)
    if not user:
        json_response(handler, 401, {"error": "Please log in first."})
        return None
    return user


def require_role(handler, role):
    user = require_user(handler)
    if not user:
        return None
    if user["role"] != role:
        json_response(handler, 403, {"error": "You do not have permission for this action."})
        return None
    return user


def save_report(upload_id, failed_rows):
    report_path = REPORT_DIR / f"reorder_report_{upload_id}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Reorder Report"
    headers = [
        "Material",
        "Material Group",
        "Material Description",
        "Current Stock",
        "Minimum Stock",
        "Reorder Quantity",
        "Latest Movement",
        "Movement Type",
        "Entry Date",
        "Reason",
    ]
    ws.append(headers)
    for row in failed_rows:
        ws.append([
            row["material"],
            row.get("material_group", ""),
            row["description"],
            row["current_stock"],
            row["minimum_stock"],
            row["reorder_quantity"],
            row["latest_quantity"],
            row["movement_type"],
            row["entry_date"],
            row["reason"],
        ])
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 42)
    wb.save(report_path)
    return report_path


def save_analysis(upload_id, failed_rows, no_criteria_count):
    analysis_path = ANALYSIS_DIR / f"analysis_{upload_id}.json"
    payload = {
        "upload_id": upload_id,
        "failed_rows": failed_rows,
        "no_criteria_count": no_criteria_count,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    write_json(analysis_path, payload)
    return analysis_path


def process_upload(file_path, original_name, category=DEFAULT_CATEGORY, uploaded_by=""):
    category = normalize_category(category)
    criteria = read_json(CRITERIA_FILE)
    materials = read_json(MATERIALS_FILE)
    uploads = read_json(UPLOADS_FILE)

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty.")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))

    idx = {name: headers.index(name) for name in REQUIRED_COLUMNS}
    upload_id = uuid.uuid4().hex[:10]
    imported_rows = 0
    received_rows = 0
    used_rows = 0
    used_quantity_total = 0
    no_criteria_count = 0
    failed_by_material = {}

    for raw in rows[1:]:
        if not raw or all(value is None for value in raw):
            continue

        material = str(raw[idx["Material"]] or "").strip()
        if not material:
            continue

        description = str(raw[idx["Material Description"]] or "").strip()
        quantity = parse_number(raw[idx["Quantity"]])
        current_stock = parse_number(raw[idx["Valuated Stock"]])
        days_between = parse_number(raw[idx["Days Between"]])
        entry_date = normalize_date(raw[idx["Entry Date"]])
        purchase_order_date = normalize_date(raw[idx["Purchase Order Date"]])
        movement_type = "Received" if quantity > 0 else "Used" if quantity < 0 else "No movement"
        if quantity > 0:
            received_rows += 1
        elif quantity < 0:
            used_rows += 1
            used_quantity_total += abs(quantity)

        materials[material] = {
            "material": material,
            "description": description,
            "current_stock": current_stock,
            "category": category,
            "material_group": material_group_for(material),
            "last_entry_date": entry_date,
            "last_quantity": quantity,
            "movement_type": movement_type,
            "days_between": days_between,
            "purchase_order_date": purchase_order_date,
        }
        imported_rows += 1

        material_criteria = get_material_criteria(criteria, category, material)
        if not material_criteria or not material_criteria.get("active", True):
            no_criteria_count += 1
            continue

        minimum_stock = parse_number(material_criteria.get("minimum_stock"))
        reorder_quantity = parse_number(material_criteria.get("reorder_quantity"))
        if current_stock <= minimum_stock:
            failed_by_material[material] = {
                "material": material,
                "material_group": material_group_for(material),
                "description": description,
                "current_stock": current_stock,
                "minimum_stock": minimum_stock,
                "reorder_quantity": reorder_quantity,
                "latest_quantity": quantity,
                "movement_type": movement_type,
                "entry_date": entry_date,
                "reason": f"Current stock {current_stock:g} is at or below minimum stock {minimum_stock:g}",
            }

    failed_rows = sorted(failed_by_material.values(), key=lambda item: (item["current_stock"], item["material"]))
    report_path = save_report(upload_id, failed_rows)
    analysis_path = save_analysis(upload_id, failed_rows, no_criteria_count)
    upload_record = {
        "id": upload_id,
        "file_name": original_name,
        "category": category,
        "category_label": CATEGORIES[category],
        "uploaded_by": uploaded_by,
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "imported_rows": imported_rows,
        "received_rows": received_rows,
        "used_rows": used_rows,
        "used_quantity_total": used_quantity_total,
        "failed_count": len(failed_rows),
        "no_criteria_count": no_criteria_count,
        "report_file": report_path.name,
        "analysis_file": analysis_path.name,
    }
    uploads.insert(0, upload_record)

    write_json(MATERIALS_FILE, materials)
    write_json(UPLOADS_FILE, uploads)

    return {
        "upload": upload_record,
        "failed_rows": failed_rows,
        "summary": build_summary(),
    }


def build_summary():
    materials = read_json(MATERIALS_FILE)
    criteria = read_json(CRITERIA_FILE)
    uploads = read_json(UPLOADS_FILE)
    group_data = get_material_group_data()
    low_stock = 0
    category_cards = {}
    for category, label in CATEGORIES.items():
        category_materials = [
            item for item in materials.values()
            if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
        ]
        category_criteria = [
            item for item in criteria.values()
            if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
        ]
        category_uploads = [
            item for item in uploads
            if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
        ]
        card_low_stock = 0
        group_cards = {
            code: {
                "code": code,
                "label": group.get("label", code),
                "source_parts_count": group.get("source_parts_count", 0),
                "materials_count": 0,
                "criteria_count": 0,
                "low_stock_count": 0,
                "active_stocks_count": 0,
            }
            for code, group in group_data.get("groups", {}).items()
        }
        for material in category_materials:
            code = str(material.get("material", ""))
            group_code = material.get("material_group") or material_group_for(code)
            if group_code:
                group_cards.setdefault(group_code, {
                    "code": group_code,
                    "label": group_code,
                    "source_parts_count": 0,
                    "materials_count": 0,
                    "criteria_count": 0,
                    "low_stock_count": 0,
                    "active_stocks_count": 0,
                })
                group_cards[group_code]["materials_count"] += 1
                if parse_number(material.get("last_quantity")) < 0:
                    group_cards[group_code]["active_stocks_count"] += 1
            settings = get_material_criteria(criteria, category, code)
            if not settings or not settings.get("active", True):
                continue
            if group_code:
                group_cards[group_code]["criteria_count"] += 1
            if parse_number(material.get("current_stock")) <= parse_number(settings.get("minimum_stock")):
                card_low_stock += 1
                if group_code:
                    group_cards[group_code]["low_stock_count"] += 1
        latest_upload = category_uploads[0] if category_uploads else None
        category_cards[category] = {
            "key": category,
            "label": label,
            "materials_count": len(category_materials),
            "criteria_count": len(category_criteria),
            "low_stock_count": card_low_stock,
            "uploads_count": len(category_uploads),
            "latest_upload": latest_upload,
            "active_stocks_count": latest_upload.get("used_rows", 0) if latest_upload else 0,
            "active_quantity_used": latest_upload.get("used_quantity_total", 0) if latest_upload else 0,
            "material_groups": group_cards,
        }
    low_stock = sum(card["low_stock_count"] for card in category_cards.values())
    return {
        "materials_count": len(materials),
        "criteria_count": len(criteria),
        "uploads_count": len(uploads),
        "low_stock_count": low_stock,
        "latest_upload": uploads[0] if uploads else None,
        "categories": category_cards,
    }


def search_materials(query, category=DEFAULT_CATEGORY, material_group=""):
    category = normalize_category(category)
    materials = read_json(MATERIALS_FILE)
    criteria = read_json(CRITERIA_FILE)
    normalized = query.strip().lower()
    results = []
    for material in materials.values():
        if normalize_category(material.get("category", DEFAULT_CATEGORY)) != category:
            continue
        code = str(material.get("material", ""))
        group_code = material.get("material_group") or material_group_for(code)
        if material_group and group_code != material_group:
            continue
        description = str(material.get("description", ""))
        haystack = f"{code} {description}".lower()
        if normalized and normalized not in haystack:
            continue
        setting = get_material_criteria(criteria, category, code) or {}
        results.append({
            "material": code,
            "description": description,
            "category": category,
            "material_group": group_code,
            "current_stock": material.get("current_stock", 0),
            "last_entry_date": material.get("last_entry_date", ""),
            "minimum_stock": setting.get("minimum_stock", ""),
            "reorder_quantity": setting.get("reorder_quantity", ""),
            "has_criteria": bool(setting),
        })
    results.sort(key=lambda item: (item["has_criteria"], item["material"]))
    return results[:25]


class AppHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/api/me":
            return json_response(self, 200, {"user": current_user(self)})
        if path == "/api/summary":
            if not require_user(self):
                return
            return json_response(self, 200, build_summary())
        if path == "/api/criteria":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            category = normalize_category(query.get("category", [DEFAULT_CATEGORY])[0])
            material_group = query.get("material_group", [""])[0]
            criteria = read_json(CRITERIA_FILE)
            filtered = {
                key: value for key, value in criteria.items()
                if normalize_category(value.get("category", DEFAULT_CATEGORY)) == category
            }
            filtered = {
                key: {**value, "material_group": material_group_for(value.get("material"))}
                for key, value in filtered.items()
            }
            if material_group:
                filtered = {
                    key: value for key, value in filtered.items()
                    if material_group_for(value.get("material")) == material_group
                }
            return json_response(self, 200, filtered)
        if path == "/api/materials":
            if not require_user(self):
                return
            parsed_query = parse_qs(parsed.query)
            query = parsed_query.get("q", [""])[0]
            category = normalize_category(parsed_query.get("category", [DEFAULT_CATEGORY])[0])
            material_group = parsed_query.get("material_group", [""])[0]
            if query:
                return json_response(self, 200, search_materials(query, category, material_group))
            rows = [
                item for item in read_json(MATERIALS_FILE).values()
                if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
                and (not material_group or (item.get("material_group") or material_group_for(item.get("material"))) == material_group)
            ]
            return json_response(self, 200, rows[:100])
        if path == "/api/material-groups":
            if not require_user(self):
                return
            group_data = get_material_group_data()
            groups = sorted(group_data.get("groups", {}).values(), key=lambda item: item.get("code", ""))
            return json_response(self, 200, {"groups": groups})
        if path == "/api/group-critical":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            category = normalize_category(query.get("category", [DEFAULT_CATEGORY])[0])
            material_group = query.get("material_group", [""])[0]
            critical_stock = query.get("critical_stock", [""])[0]
            return json_response(self, 200, analyze_group_critical(category, material_group, critical_stock))
        if path == "/api/uploads":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            category = query.get("category", [""])[0]
            uploads = read_json(UPLOADS_FILE)
            if category:
                category = normalize_category(category)
                uploads = [
                    item for item in uploads
                    if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
                ]
            return json_response(self, 200, uploads)
        if path == "/api/analysis":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            upload_id = query.get("upload_id", [""])[0]
            category = query.get("category", [""])[0]
            material_group = query.get("material_group", [""])[0]
            uploads = read_json(UPLOADS_FILE)
            if category:
                category = normalize_category(category)
                uploads = [
                    item for item in uploads
                    if normalize_category(item.get("category", DEFAULT_CATEGORY)) == category
                ]
            upload = next((item for item in uploads if item.get("id") == upload_id), uploads[0] if uploads else None)
            if not upload:
                return json_response(self, 200, {"upload": None, "failed_rows": [], "no_criteria_count": 0})
            analysis_file = upload.get("analysis_file") or f"analysis_{upload['id']}.json"
            analysis_path = ANALYSIS_DIR / analysis_file
            if not analysis_path.exists():
                return json_response(self, 200, {"upload": upload, "failed_rows": [], "no_criteria_count": 0})
            analysis = read_json(analysis_path)
            failed_rows = analysis.get("failed_rows", [])
            if material_group:
                failed_rows = [
                    row for row in failed_rows
                    if material_group_for(row.get("material")) == material_group
                ]
            failed_rows = [
                {**row, "material_group": row.get("material_group") or material_group_for(row.get("material"))}
                for row in failed_rows
            ]
            return json_response(self, 200, {
                "upload": upload,
                "failed_rows": failed_rows,
                "no_criteria_count": analysis.get("no_criteria_count", 0),
                "generated_at": analysis.get("generated_at"),
            })
        if path == "/api/download":
            if not require_user(self):
                return
            query = parse_qs(parsed.query)
            name = query.get("file", [""])[0]
            report_path = (REPORT_DIR / name).resolve()
            if REPORT_DIR.resolve() not in report_path.parents or not report_path.exists():
                return json_response(self, 404, {"error": "Report not found."})
            body = report_path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{report_path.name}"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        return self.serve_static(path)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/login":
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            email = str(payload.get("email", "")).strip().lower()
            password = str(payload.get("password", ""))
            user = DEMO_USERS.get(email)
            if not user or user["password"] != password:
                return json_response(self, 401, {"error": "Invalid email or password."})
            session_id = uuid.uuid4().hex
            SESSIONS[session_id] = email
            body = json.dumps({"user": {key: value for key, value in user.items() if key != "password"}}).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Set-Cookie", f"session_id={session_id}; Path=/; SameSite=Lax")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/logout":
            cookies = parse_cookies(self.headers.get("Cookie"))
            session_id = cookies.get("session_id")
            if session_id in SESSIONS:
                del SESSIONS[session_id]
            body = json.dumps({"ok": True}).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Set-Cookie", "session_id=; Path=/; Max-Age=0; SameSite=Lax")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/criteria":
            if not require_role(self, "admin"):
                return
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            material = str(payload.get("material", "")).strip()
            category = normalize_category(payload.get("category", DEFAULT_CATEGORY))
            if not material:
                return json_response(self, 400, {"error": "Material code is required."})
            criteria = read_json(CRITERIA_FILE)
            key = criteria_key(category, material)
            for existing_key, existing_value in list(criteria.items()):
                if existing_key != key and criteria_matches(existing_value, category, material):
                    del criteria[existing_key]
            criteria[key] = {
                "material": material,
                "description": str(payload.get("description", "")).strip(),
                "category": category,
                "minimum_stock": parse_number(payload.get("minimum_stock")),
                "reorder_quantity": parse_number(payload.get("reorder_quantity")),
                "active": bool(payload.get("active", True)),
            }
            write_json(CRITERIA_FILE, criteria)
            return json_response(self, 200, {"criteria": criteria[key], "summary": build_summary()})

        if parsed.path == "/api/group-critical":
            if not require_role(self, "admin"):
                return
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            category = normalize_category(payload.get("category", DEFAULT_CATEGORY))
            material_group = str(payload.get("material_group", "")).strip()
            if not material_group:
                return json_response(self, 400, {"error": "Please select a material group first."})
            group_settings = read_json(GROUP_CRITERIA_FILE)
            key = group_criteria_key(category, material_group)
            group_settings[key] = {
                "category": category,
                "material_group": material_group,
                "critical_stock": parse_number(payload.get("critical_stock")),
                "reorder_quantity": parse_number(payload.get("reorder_quantity")),
                "active": bool(payload.get("active", True)),
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            write_json(GROUP_CRITERIA_FILE, group_settings)
            analysis = analyze_group_critical(category, material_group, group_settings[key]["critical_stock"])
            return json_response(self, 200, {"criteria": group_settings[key], "analysis": analysis})

        if parsed.path == "/api/upload":
            if not require_user(self):
                return
            form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            })
            file_item = form["file"] if "file" in form else None
            if not file_item or not file_item.filename:
                return json_response(self, 400, {"error": "Please choose an Excel file."})
            category = normalize_category(form.getvalue("category", DEFAULT_CATEGORY))
            upload_name = Path(file_item.filename).name
            saved_path = UPLOAD_DIR / f"{uuid.uuid4().hex}_{upload_name}"
            with saved_path.open("wb") as output:
                shutil.copyfileobj(file_item.file, output)
            try:
                user = current_user(self) or {}
                result = process_upload(saved_path, upload_name, category, user.get("email", ""))
                return json_response(self, 200, result)
            except Exception as exc:
                return json_response(self, 400, {"error": str(exc)})

        return json_response(self, 404, {"error": "Unknown endpoint."})

    def serve_static(self, path):
        if path == "/":
            path = "/index.html"
        file_path = (STATIC_DIR / path.lstrip("/")).resolve()
        if STATIC_DIR.resolve() not in file_path.parents and file_path != STATIC_DIR.resolve():
            return json_response(self, 403, {"error": "Forbidden."})
        if not file_path.exists() or not file_path.is_file():
            return json_response(self, 404, {"error": "Not found."})
        body = file_path.read_bytes()
        content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        sys.stderr.write("%s - %s\n" % (self.address_string(), format % args))


if __name__ == "__main__":
    ensure_dirs()
    server = ThreadingHTTPServer(("localhost", 8000), AppHandler)
    print("Tata Inventory Criteria Checker running at http://localhost:8000")
    server.serve_forever()
